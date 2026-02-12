"""
02_05_reports_cross_market.py - Генерація звітів для кожного ринку

Phase 1, Step 5: Reports Generation

Вхідні дані:
- data/processed_data/01_per_market/{CLIENT_ID}/03_did_analysis_{CLIENT_ID}/_stats/drugs_summary_{CLIENT_ID}.csv
- data/processed_data/01_per_market/{CLIENT_ID}/03_did_analysis_{CLIENT_ID}/did_results_{CLIENT_ID}.csv
- data/processed_data/01_per_market/{CLIENT_ID}/04_substitute_shares_{CLIENT_ID}/substitute_shares_{CLIENT_ID}.csv

Вихідні дані:
- results/data_reports/reports_{CLIENT_ID}/01_technical_report_{CLIENT_ID}.xlsx
- results/data_reports/reports_{CLIENT_ID}/02_business_report_{CLIENT_ID}.xlsx
- results/cross_market_data/cross_market_{CLIENT_ID}.csv

Використання:
    python exec_scripts/01_did_processing/02_05_reports_cross_market.py --market_id 28670
    python exec_scripts/01_did_processing/02_05_reports_cross_market.py --all
"""

import pandas as pd
import numpy as np
import os
import sys
import argparse
from datetime import datetime
from typing import List, Dict, Any, Optional, Tuple
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter

# Додаємо шлях до project_core
sys.path.insert(0, os.path.join(os.path.dirname(__file__), '..', '..'))
from project_core.data_config.paths_config import (
    PROCESSED_DATA_PATH,
    RESULTS_PATH,
    get_market_paths,
    load_target_pharmacies
)


# ============================================================================
# КОНФІГУРАЦІЯ КОЛОНОК ЗВІТІВ
# ============================================================================

# Колонки препарату (заповнюються тільки в першому рядку)
DRUG_COLUMNS = [
    ('DRUGS_ID', 'ID препарату'),
    ('DRUGS_NAME', 'Назва препарату'),
    ('INN_ID', 'ID МНН групи'),
    ('INN_NAME', 'Назва МНН групи'),
    ('NFC1_ID', 'Широка категорія форми (NFC1)'),
    ('EVENTS_COUNT', 'Кількість stock-out подій'),
    ('TOTAL_STOCKOUT_WEEKS', 'Загальна тривалість stock-out (тижні)'),
    ('FIRST_STOCKOUT_DATE', 'Дата першого stock-out'),
    ('LAST_STOCKOUT_DATE', 'Дата останнього stock-out'),
    ('INTERNAL_LIFT', 'Загальний внутрішній ліфт (упак.)'),
    ('LOST_SALES', 'Загальні втрачені продажі (упак.)'),
    ('TOTAL_EFFECT', 'Загально перерозподілено (упак.)'),
    ('TOTAL_LIFT_SAME_NFC1', 'Ліфт тієї ж категорії (упак.)'),
    ('TOTAL_LIFT_DIFF_NFC1', 'Ліфт іншої категорії (упак.)'),
    ('SHARE_INTERNAL', 'Частка внутрішньої субституції (%)'),
    ('SHARE_LOST', 'Частка втрат до конкурентів (%)'),
    ('SHARE_SAME_NFC1', 'Частка тієї ж категорії (%)'),
    ('SHARE_DIFF_NFC1', 'Частка іншої категорії (%)'),
]

# Колонки субститутів (та ж категорія)
SAME_NFC1_COLUMNS = [
    ('SAME_NFC1_DRUG_NAME', 'Препарати-субститути (та ж категорія)'),
    ('SAME_NFC1_DRUG_ID', 'ID субститутів (та ж категорія)'),
    ('SAME_NFC1_SUBSTITUTE_SHARE', 'Відсоток заміщення (та ж категорія)'),
]

# Колонки субститутів (інша категорія)
DIFF_NFC1_COLUMNS = [
    ('DIFF_NFC1_DRUG_NAME', 'Препарати-субститути (інша категорія)'),
    ('DIFF_NFC1_DRUG_ID', 'ID субститутів (інша категорія)'),
    ('DIFF_NFC1_SUBSTITUTE_SHARE', 'Відсоток заміщення (інша категорія)'),
    ('DIFF_NFC1_ID', 'Форма випуску (інша категорія)'),
]

# Колонки класифікації
CLASSIFICATION_COLUMNS = [
    ('CLASSIFICATION', 'Категорія'),
    ('RECOMMENDATION', 'Рекомендація'),
]

# Всі колонки технічного звіту
ALL_TECH_COLUMNS = DRUG_COLUMNS + SAME_NFC1_COLUMNS + DIFF_NFC1_COLUMNS + CLASSIFICATION_COLUMNS

# Колонки бізнес-звіту (скорочена версія)
BUSINESS_DRUG_COLUMNS = [
    ('DRUGS_ID', 'ID препарату'),
    ('DRUGS_NAME', 'Назва препарату'),
    ('INN_ID', 'ID МНН групи'),
    ('INN_NAME', 'Назва МНН групи'),
    ('NFC1_ID', 'Широка категорія форми (NFC1)'),
    ('SHARE_INTERNAL', 'Частка внутрішньої субституції (%)'),
    ('SHARE_LOST', 'Частка втрат до конкурентів (%)'),
    ('SHARE_SAME_NFC1', 'Частка тієї ж категорії (%)'),
    ('SHARE_DIFF_NFC1', 'Частка іншої категорії (%)'),
]

ALL_BUSINESS_COLUMNS = BUSINESS_DRUG_COLUMNS + SAME_NFC1_COLUMNS + DIFF_NFC1_COLUMNS + CLASSIFICATION_COLUMNS


# ============================================================================
# СТИЛІ EXCEL
# ============================================================================

HEADER_FONT_TECH = Font(bold=True, size=9, color='666666')
HEADER_FONT_HUMAN = Font(bold=True, size=10)
HEADER_FILL = PatternFill(start_color='E8E8E8', end_color='E8E8E8', fill_type='solid')
HEADER_ALIGNMENT = Alignment(horizontal='center', vertical='center', wrap_text=True)

CATEGORY_COLORS = {
    'CRITICAL': PatternFill(start_color='FFCCCC', end_color='FFCCCC', fill_type='solid'),
    'SUBSTITUTABLE': PatternFill(start_color='CCFFCC', end_color='CCFFCC', fill_type='solid'),
    'MIXED': PatternFill(start_color='FFFFCC', end_color='FFFFCC', fill_type='solid')
}

THIN_BORDER = Border(
    left=Side(style='thin'),
    right=Side(style='thin'),
    top=Side(style='thin'),
    bottom=Side(style='thin')
)

# Ширина колонок
COLUMN_WIDTHS = {
    'DRUGS_ID': 12,
    'DRUGS_NAME': 60,
    'INN_ID': 10,
    'INN_NAME': 15,
    'NFC1_ID': 30,
    'EVENTS_COUNT': 12,
    'TOTAL_STOCKOUT_WEEKS': 12,
    'FIRST_STOCKOUT_DATE': 14,
    'LAST_STOCKOUT_DATE': 14,
    'INTERNAL_LIFT': 15,
    'LOST_SALES': 15,
    'TOTAL_EFFECT': 15,
    'TOTAL_LIFT_SAME_NFC1': 15,
    'TOTAL_LIFT_DIFF_NFC1': 15,
    'SHARE_INTERNAL': 15,
    'SHARE_LOST': 15,
    'SHARE_SAME_NFC1': 15,
    'SHARE_DIFF_NFC1': 15,
    'SAME_NFC1_DRUG_NAME': 60,
    'SAME_NFC1_DRUG_ID': 12,
    'SAME_NFC1_SUBSTITUTE_SHARE': 15,
    'DIFF_NFC1_DRUG_NAME': 60,
    'DIFF_NFC1_DRUG_ID': 12,
    'DIFF_NFC1_SUBSTITUTE_SHARE': 15,
    'DIFF_NFC1_ID': 30,
    'CLASSIFICATION': 15,
    'RECOMMENDATION': 35,
}

# Колонки з відсотками
PERCENT_COLUMNS = [
    'SHARE_INTERNAL', 'SHARE_LOST', 'SHARE_SAME_NFC1', 'SHARE_DIFF_NFC1',
    'SAME_NFC1_SUBSTITUTE_SHARE', 'DIFF_NFC1_SUBSTITUTE_SHARE'
]

# Колонки з float
FLOAT_COLUMNS = [
    'INTERNAL_LIFT', 'LOST_SALES', 'TOTAL_EFFECT',
    'TOTAL_LIFT_SAME_NFC1', 'TOTAL_LIFT_DIFF_NFC1'
]


# ============================================================================
# ДОПОМІЖНІ ФУНКЦІЇ
# ============================================================================

def get_recommendation(classification: str) -> str:
    """Отримати рекомендацію на основі класифікації."""
    if classification == 'CRITICAL':
        return 'KEEP - High loss to competitors'
    elif classification == 'SUBSTITUTABLE':
        return 'CONSIDER_REMOVAL - Good internal substitution'
    return 'ANALYZE - Mixed results'


def get_substitutes_for_drug(
    substitute_shares: pd.DataFrame,
    drug_id: int,
    same_nfc1: bool = True
) -> List[Dict[str, Any]]:
    """
    Отримати субститути для препарату, відсортовані за SUBSTITUTE_SHARE спадаючи.
    """
    drug_subs = substitute_shares[
        (substitute_shares['STOCKOUT_DRUG_ID'] == drug_id) &
        (substitute_shares['SAME_NFC1'] == same_nfc1)
    ].sort_values('SUBSTITUTE_SHARE', ascending=False)

    result = []
    for _, row in drug_subs.iterrows():
        result.append({
            'DRUG_NAME': row['SUBSTITUTE_DRUG_NAME'],
            'DRUG_ID': row['SUBSTITUTE_DRUG_ID'],
            'SUBSTITUTE_SHARE': row['SUBSTITUTE_SHARE'],
            'NFC1_ID': row['SUBSTITUTE_NFC1_ID'] if not same_nfc1 else None
        })
    return result


def load_market_data(client_id: int) -> Tuple[pd.DataFrame, pd.DataFrame, pd.DataFrame]:
    """
    Завантажити всі необхідні дані для ринку.

    Returns:
        Tuple of (drugs_summary, did_results, substitute_shares)
    """
    paths = get_market_paths(client_id)
    did_path = paths['did_analysis']
    substitute_path = paths['substitute_shares']

    # Шляхи до файлів
    drugs_summary_file = did_path / '_stats' / f'drugs_summary_{client_id}.csv'
    did_results_file = did_path / f'did_results_{client_id}.csv'
    substitute_shares_file = substitute_path / f'substitute_shares_{client_id}.csv'

    # Перевірка існування файлів
    for filepath, name in [
        (drugs_summary_file, 'drugs_summary'),
        (did_results_file, 'did_results'),
        (substitute_shares_file, 'substitute_shares')
    ]:
        if not filepath.exists():
            raise FileNotFoundError(f"{name} file not found: {filepath}")

    drugs_summary = pd.read_csv(drugs_summary_file)
    did_results = pd.read_csv(did_results_file)
    substitute_shares = pd.read_csv(substitute_shares_file)

    return drugs_summary, did_results, substitute_shares


def prepare_base_dataframe(
    drugs_summary: pd.DataFrame,
    did_results: pd.DataFrame,
    substitute_shares: pd.DataFrame
) -> pd.DataFrame:
    """
    Підготувати базовий DataFrame з усіма агрегованими даними.
    """
    # Агрегація дат stock-out з did_results
    date_agg = did_results.groupby('DRUGS_ID').agg({
        'STOCKOUT_START': 'min',
        'STOCKOUT_END': 'max',
        'STOCKOUT_WEEKS': 'sum'
    }).reset_index()
    date_agg.columns = ['DRUGS_ID', 'FIRST_STOCKOUT_DATE', 'LAST_STOCKOUT_DATE', 'TOTAL_STOCKOUT_WEEKS']

    # Агрегація ліфтів з substitute_shares
    lift_agg = substitute_shares.groupby('STOCKOUT_DRUG_ID').agg({
        'LIFT_SAME_NFC1': 'sum',
        'LIFT_DIFF_NFC1': 'sum',
    }).reset_index()
    lift_agg.columns = ['DRUGS_ID', 'TOTAL_LIFT_SAME_NFC1', 'TOTAL_LIFT_DIFF_NFC1']

    # Об'єднуємо всі дані
    base_df = drugs_summary.copy()
    base_df = base_df.merge(date_agg, on='DRUGS_ID', how='left')
    base_df = base_df.merge(lift_agg, on='DRUGS_ID', how='left')

    # Заповнюємо NaN
    base_df['TOTAL_LIFT_SAME_NFC1'] = base_df['TOTAL_LIFT_SAME_NFC1'].fillna(0)
    base_df['TOTAL_LIFT_DIFF_NFC1'] = base_df['TOTAL_LIFT_DIFF_NFC1'].fillna(0)
    base_df['TOTAL_STOCKOUT_WEEKS'] = base_df['TOTAL_STOCKOUT_WEEKS'].fillna(0)

    # Перераховуємо SHARE_SAME_NFC1 та SHARE_DIFF_NFC1 як ratio сум (не mean)
    # Це дає "зважене середнє" - події з більшим LIFT мають більший вплив
    # Консистентно з оригінальною реалізацією та правильно для Phase 2 aggregation
    base_df['SHARE_SAME_NFC1'] = np.where(
        base_df['INTERNAL_LIFT'] > 0,
        base_df['TOTAL_LIFT_SAME_NFC1'] / base_df['INTERNAL_LIFT'],
        0.0
    )
    base_df['SHARE_DIFF_NFC1'] = np.where(
        base_df['INTERNAL_LIFT'] > 0,
        base_df['TOTAL_LIFT_DIFF_NFC1'] / base_df['INTERNAL_LIFT'],
        0.0
    )

    # Додаємо рекомендації
    base_df['RECOMMENDATION'] = base_df['CLASSIFICATION'].apply(get_recommendation)

    return base_df


def build_report_rows(
    base_df: pd.DataFrame,
    substitute_shares: pd.DataFrame,
    columns: List[Tuple[str, str]]
) -> pd.DataFrame:
    """
    Сформувати рядки звіту з вертикальним списком субститутів.
    """
    # Сортуємо препарати: CRITICAL спочатку, потім по SHARE_LOST
    category_order = {'CRITICAL': 0, 'MIXED': 1, 'SUBSTITUTABLE': 2}
    sorted_drugs = base_df.copy()
    sorted_drugs['_sort'] = sorted_drugs['CLASSIFICATION'].map(category_order).fillna(1)
    sorted_drugs = sorted_drugs.sort_values(
        ['_sort', 'SHARE_LOST'],
        ascending=[True, False]
    ).drop('_sort', axis=1)

    # Визначаємо які колонки потрібні
    tech_headers = [col[0] for col in columns]
    drug_cols = [col[0] for col in DRUG_COLUMNS if col[0] in tech_headers]
    same_nfc1_cols = [col[0] for col in SAME_NFC1_COLUMNS if col[0] in tech_headers]
    diff_nfc1_cols = [col[0] for col in DIFF_NFC1_COLUMNS if col[0] in tech_headers]
    class_cols = [col[0] for col in CLASSIFICATION_COLUMNS if col[0] in tech_headers]

    report_rows = []

    for _, drug in sorted_drugs.iterrows():
        drug_id = drug['DRUGS_ID']

        # Отримати субститути
        same_nfc1_subs = get_substitutes_for_drug(substitute_shares, drug_id, same_nfc1=True)
        diff_nfc1_subs = get_substitutes_for_drug(substitute_shares, drug_id, same_nfc1=False)

        # Максимальна кількість рядків для цього препарату
        max_rows = max(1, len(same_nfc1_subs), len(diff_nfc1_subs))

        for row_idx in range(max_rows):
            row = {}

            # Колонки препарату - тільки в першому рядку
            if row_idx == 0:
                for col in drug_cols:
                    if col in drug.index:
                        row[col] = drug[col]
                    else:
                        row[col] = ''

                # Колонки класифікації - тільки в першому рядку
                for col in class_cols:
                    if col in drug.index:
                        row[col] = drug[col]
                    else:
                        row[col] = ''
            else:
                # Порожні колонки для наступних рядків
                for col in drug_cols:
                    row[col] = ''
                for col in class_cols:
                    row[col] = ''

            # Колонки субститутів (та ж категорія)
            if row_idx < len(same_nfc1_subs) and same_nfc1_cols:
                sub = same_nfc1_subs[row_idx]
                if 'SAME_NFC1_DRUG_NAME' in same_nfc1_cols:
                    row['SAME_NFC1_DRUG_NAME'] = sub['DRUG_NAME']
                if 'SAME_NFC1_DRUG_ID' in same_nfc1_cols:
                    row['SAME_NFC1_DRUG_ID'] = sub['DRUG_ID']
                if 'SAME_NFC1_SUBSTITUTE_SHARE' in same_nfc1_cols:
                    row['SAME_NFC1_SUBSTITUTE_SHARE'] = sub['SUBSTITUTE_SHARE']
            else:
                for col in same_nfc1_cols:
                    row[col] = ''

            # Колонки субститутів (інша категорія)
            if row_idx < len(diff_nfc1_subs) and diff_nfc1_cols:
                sub = diff_nfc1_subs[row_idx]
                if 'DIFF_NFC1_DRUG_NAME' in diff_nfc1_cols:
                    row['DIFF_NFC1_DRUG_NAME'] = sub['DRUG_NAME']
                if 'DIFF_NFC1_DRUG_ID' in diff_nfc1_cols:
                    row['DIFF_NFC1_DRUG_ID'] = sub['DRUG_ID']
                if 'DIFF_NFC1_SUBSTITUTE_SHARE' in diff_nfc1_cols:
                    row['DIFF_NFC1_SUBSTITUTE_SHARE'] = sub['SUBSTITUTE_SHARE']
                if 'DIFF_NFC1_ID' in diff_nfc1_cols:
                    row['DIFF_NFC1_ID'] = sub['NFC1_ID']
            else:
                for col in diff_nfc1_cols:
                    row[col] = ''

            report_rows.append(row)

    report_df = pd.DataFrame(report_rows)
    # Впорядкувати колонки
    report_df = report_df[tech_headers]

    return report_df


def create_excel_report(
    report_df: pd.DataFrame,
    columns: List[Tuple[str, str]],
    output_path: str,
    sheet_name: str = 'Report'
) -> None:
    """
    Створити Excel-файл зі звітом.
    """
    tech_headers = [col[0] for col in columns]
    human_headers = [col[1] for col in columns]

    # Створення workbook
    wb = Workbook()
    ws = wb.active
    ws.title = sheet_name

    # Рядок 1: Технічні заголовки
    for col_idx, header in enumerate(tech_headers, 1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.font = HEADER_FONT_TECH
        cell.fill = HEADER_FILL
        cell.alignment = HEADER_ALIGNMENT
        cell.border = THIN_BORDER

    # Рядок 2: Людські заголовки
    for col_idx, header in enumerate(human_headers, 1):
        cell = ws.cell(row=2, column=col_idx, value=header)
        cell.font = HEADER_FONT_HUMAN
        cell.fill = HEADER_FILL
        cell.alignment = HEADER_ALIGNMENT
        cell.border = THIN_BORDER

    # Дані
    current_category = None
    for row_idx, row_data in enumerate(report_df.values, 3):
        for col_idx, value in enumerate(row_data, 1):
            cell = ws.cell(row=row_idx, column=col_idx)
            col_name = tech_headers[col_idx - 1]

            # Форматування значень
            if col_name in PERCENT_COLUMNS:
                if value != '' and pd.notna(value):
                    try:
                        val = float(value)
                        # SUBSTITUTE_SHARE вже в відсотках (напр. 9.29)
                        if col_name in ['SAME_NFC1_SUBSTITUTE_SHARE', 'DIFF_NFC1_SUBSTITUTE_SHARE']:
                            cell.value = val / 100  # Конвертуємо в частку для Excel формату
                        else:
                            cell.value = val
                        cell.number_format = '0.00%'
                    except (ValueError, TypeError):
                        cell.value = value
                else:
                    cell.value = ''
            elif col_name in FLOAT_COLUMNS:
                if value != '' and pd.notna(value):
                    try:
                        cell.value = round(float(value), 2)
                    except (ValueError, TypeError):
                        cell.value = value
                else:
                    cell.value = ''
            else:
                cell.value = value if pd.notna(value) else ''

            cell.border = THIN_BORDER

            # Відслідковуємо категорію для забарвлення
            if col_name == 'CLASSIFICATION' and value != '' and pd.notna(value):
                current_category = value

        # Застосувати колір до всього рядка на основі поточної категорії
        if current_category in CATEGORY_COLORS:
            for c in range(1, len(tech_headers) + 1):
                ws.cell(row=row_idx, column=c).fill = CATEGORY_COLORS[current_category]

    # Ширина колонок
    for col_idx, col_name in enumerate(tech_headers, 1):
        ws.column_dimensions[get_column_letter(col_idx)].width = COLUMN_WIDTHS.get(col_name, 12)

    # Фіксувати заголовки
    ws.freeze_panes = 'A3'

    # Висота рядків заголовків
    ws.row_dimensions[1].height = 20
    ws.row_dimensions[2].height = 40

    # Зберегти
    wb.save(output_path)


def create_cross_market_csv(
    base_df: pd.DataFrame,
    substitute_shares: pd.DataFrame,
    client_id: int,
    output_path: str
) -> None:
    """
    Створити CSV-файл для cross-market аналізу.

    Формат: flat data з усіма ключовими метриками.
    """
    # Додаємо CLIENT_ID до base_df
    cross_df = base_df.copy()
    cross_df.insert(0, 'CLIENT_ID', client_id)

    # Сортуємо по CLASSIFICATION і SHARE_LOST
    category_order = {'CRITICAL': 0, 'MIXED': 1, 'SUBSTITUTABLE': 2}
    cross_df['_sort'] = cross_df['CLASSIFICATION'].map(category_order).fillna(1)
    cross_df = cross_df.sort_values(
        ['_sort', 'SHARE_LOST'],
        ascending=[True, False]
    ).drop('_sort', axis=1)

    # Вибираємо колонки для CSV
    csv_columns = [
        'CLIENT_ID', 'DRUGS_ID', 'DRUGS_NAME', 'INN_ID', 'INN_NAME', 'NFC1_ID',
        'EVENTS_COUNT', 'TOTAL_STOCKOUT_WEEKS', 'FIRST_STOCKOUT_DATE', 'LAST_STOCKOUT_DATE',
        'INTERNAL_LIFT', 'LOST_SALES', 'TOTAL_EFFECT',
        'TOTAL_LIFT_SAME_NFC1', 'TOTAL_LIFT_DIFF_NFC1',
        'SHARE_INTERNAL', 'SHARE_LOST', 'SHARE_SAME_NFC1', 'SHARE_DIFF_NFC1',
        'CLASSIFICATION', 'RECOMMENDATION'
    ]

    # Залишаємо тільки наявні колонки
    available_columns = [col for col in csv_columns if col in cross_df.columns]
    cross_df = cross_df[available_columns]

    # Зберігаємо
    cross_df.to_csv(output_path, index=False)


# ============================================================================
# ОСНОВНА ФУНКЦІЯ ОБРОБКИ
# ============================================================================

def process_market(client_id: int) -> Dict[str, Any]:
    """
    Обробити один ринок: створити всі звіти.

    Returns:
        Dict зі статистикою обробки
    """
    start_time = datetime.now()
    print(f"\n{'='*60}")
    print(f"Processing market: {client_id}")
    print(f"{'='*60}")

    # Створюємо директорії
    reports_dir = RESULTS_PATH / 'data_reports' / f'reports_{client_id}'
    cross_market_dir = RESULTS_PATH / 'cross_market_data'

    reports_dir.mkdir(parents=True, exist_ok=True)
    cross_market_dir.mkdir(parents=True, exist_ok=True)

    # Завантажуємо дані
    print("Loading data...")
    drugs_summary, did_results, substitute_shares = load_market_data(client_id)
    print(f"  drugs_summary: {len(drugs_summary)} препаратів")
    print(f"  did_results: {len(did_results)} подій")
    print(f"  substitute_shares: {len(substitute_shares)} пар")

    # Підготовка базового DataFrame
    print("Preparing base dataframe...")
    base_df = prepare_base_dataframe(drugs_summary, did_results, substitute_shares)

    # Статистика по категоріях
    category_counts = base_df['CLASSIFICATION'].value_counts()
    print(f"\nCategory distribution:")
    for cat, count in category_counts.items():
        print(f"  {cat}: {count}")

    # 1. Технічний звіт
    print("\nGenerating technical report...")
    tech_report_df = build_report_rows(base_df, substitute_shares, ALL_TECH_COLUMNS)
    tech_output = reports_dir / f'01_technical_report_{client_id}.xlsx'
    create_excel_report(tech_report_df, ALL_TECH_COLUMNS, str(tech_output), 'Technical Report')
    print(f"  Saved: {tech_output}")
    print(f"  Rows: {len(tech_report_df)}")

    # 2. Бізнес-звіт
    print("\nGenerating business report...")
    business_report_df = build_report_rows(base_df, substitute_shares, ALL_BUSINESS_COLUMNS)
    business_output = reports_dir / f'02_business_report_{client_id}.xlsx'
    create_excel_report(business_report_df, ALL_BUSINESS_COLUMNS, str(business_output), 'Business Report')
    print(f"  Saved: {business_output}")
    print(f"  Rows: {len(business_report_df)}")

    # 3. Cross-market CSV
    print("\nGenerating cross-market CSV...")
    csv_output = cross_market_dir / f'cross_market_{client_id}.csv'
    create_cross_market_csv(base_df, substitute_shares, client_id, str(csv_output))
    print(f"  Saved: {csv_output}")
    print(f"  Rows: {len(base_df)}")

    # Статистика
    end_time = datetime.now()
    processing_time = (end_time - start_time).total_seconds()

    result = {
        'client_id': client_id,
        'drugs_count': len(drugs_summary),
        'events_count': len(did_results),
        'pairs_count': len(substitute_shares),
        'tech_report_rows': len(tech_report_df),
        'business_report_rows': len(business_report_df),
        'cross_market_rows': len(base_df),
        'critical_count': category_counts.get('CRITICAL', 0),
        'substitutable_count': category_counts.get('SUBSTITUTABLE', 0),
        'processing_time_sec': processing_time
    }

    print(f"\n[OK] Market {client_id} processed in {processing_time:.2f}s")

    return result


# ============================================================================
# ТОЧКА ВХОДУ
# ============================================================================

def main():
    parser = argparse.ArgumentParser(
        description='Generate reports for market analysis'
    )
    parser.add_argument(
        '--market_id',
        type=int,
        help='Process single market by CLIENT_ID'
    )
    parser.add_argument(
        '--all',
        action='store_true',
        help='Process all markets'
    )

    args = parser.parse_args()

    if not args.market_id and not args.all:
        parser.error("Either --market_id or --all must be specified")

    print("="*60)
    print("REPORT GENERATION - Phase 1, Step 5")
    print("="*60)
    print(f"Started: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}")

    # Визначаємо список ринків для обробки
    if args.all:
        markets = load_target_pharmacies()
        print(f"Processing ALL markets: {markets}")
    else:
        markets = [args.market_id]
        print(f"Processing single market: {args.market_id}")

    # Обробляємо кожен ринок
    results = []
    for market_id in markets:
        try:
            result = process_market(market_id)
            results.append(result)
        except FileNotFoundError as e:
            print(f"\n[ERROR] Market {market_id}: {e}")
            continue
        except Exception as e:
            print(f"\n[ERROR] Market {market_id}: {e}")
            import traceback
            traceback.print_exc()
            continue

    # Підсумок
    print("\n" + "="*60)
    print("SUMMARY")
    print("="*60)

    if results:
        total_drugs = sum(r['drugs_count'] for r in results)
        total_events = sum(r['events_count'] for r in results)
        total_pairs = sum(r['pairs_count'] for r in results)
        total_critical = sum(r['critical_count'] for r in results)
        total_substitutable = sum(r['substitutable_count'] for r in results)
        total_time = sum(r['processing_time_sec'] for r in results)

        print(f"Markets processed: {len(results)}")
        print(f"Total drugs: {total_drugs}")
        print(f"Total events: {total_events}")
        print(f"Total substitute pairs: {total_pairs}")
        print(f"Total CRITICAL: {total_critical}")
        print(f"Total SUBSTITUTABLE: {total_substitutable}")
        print(f"Total processing time: {total_time:.2f}s")

        print("\nPer-market summary:")
        for r in results:
            print(f"  {r['client_id']}: {r['drugs_count']} drugs, "
                  f"{r['critical_count']} CRITICAL, {r['substitutable_count']} SUBSTITUTABLE")
    else:
        print("No markets were processed successfully")

    print(f"\nCompleted: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}")


if __name__ == '__main__':
    main()
