"""
01_data_preparation.py - Підготовка даних для Phase 2 Cross-Market Aggregation

Phase 2, Step 1: Data Preparation

Вхідні дані:
- data/raw/Rd2_{CLIENT_ID}.csv (для all_drugs_list)
- results/cross_market_data/cross_market_{CLIENT_ID}.csv (Phase 1 результати)
- data/processed_data/00_preproc_results/target_pharmacies_list.csv

Вихідні дані:
- results/substitution_research/01_preparation/all_drugs_list.csv
- results/substitution_research/01_preparation/researched_drugs_list.csv
- results/substitution_research/01_preparation/researched_drugs_coefficients.csv
- results/substitution_research/01_preparation/coverage_analysis.csv
- results/substitution_research/01_preparation/validation_report.txt
- results/substitution_research/01_preparation/prep_business_reports/*.xlsx

Використання:
    python exec_scripts/02_substitution_coefficients/01_data_preparation.py
"""

import pandas as pd
import numpy as np
import os
import sys
from pathlib import Path
from datetime import datetime
from typing import List, Dict, Tuple, Optional
from collections import defaultdict

# Додаємо шлях до project_core
SCRIPT_PATH = Path(__file__).resolve()
PROJECT_ROOT = SCRIPT_PATH.parent.parent.parent
sys.path.insert(0, str(PROJECT_ROOT))

from project_core.data_config.paths_config import (
    RAW_DATA_PATH,
    PROCESSED_DATA_PATH,
    RESULTS_PATH,
    load_target_pharmacies
)
from project_core.sub_coef_config.coverage_thresholds import (
    COVERAGE_HIGH,
    COVERAGE_MEDIUM,
    COVERAGE_LOW,
    get_coverage_cluster
)


# =============================================================================
# КОНСТАНТИ
# =============================================================================

# Шляхи виходу
OUTPUT_BASE_PATH = RESULTS_PATH / "substitution_research" / "01_preparation"
OUTPUT_BUSINESS_PATH = OUTPUT_BASE_PATH / "prep_business_reports"

# Вхідні шляхи
CROSS_MARKET_PATH = RESULTS_PATH / "cross_market_data"

# Колонки для читання з raw файлів (оригінальні назви)
RAW_COLUMNS_ORIGINAL = ['DRUGS_ID', 'Full medication name', 'INN_ID', 'INN']

# Перейменування колонок з raw
RAW_COLUMN_RENAME = {
    'Full medication name': 'DRUGS_NAME',
    'INN': 'INN_NAME'
}

# Колонки для читання з cross_market файлів
CROSS_MARKET_COLUMNS = [
    'CLIENT_ID', 'DRUGS_ID', 'DRUGS_NAME', 'INN_ID', 'INN_NAME', 'NFC1_ID',
    'INTERNAL_LIFT', 'SHARE_INTERNAL', 'EVENTS_COUNT'
]


# =============================================================================
# ЗАВАНТАЖЕННЯ ДАНИХ
# =============================================================================

def load_all_drugs_from_raw() -> pd.DataFrame:
    """
    Завантажити список всіх унікальних препаратів з raw файлів.

    Returns:
        DataFrame з колонками: DRUGS_ID, DRUGS_NAME, INN_ID, INN_NAME
    """
    print("\n" + "=" * 60)
    print("ЗАВАНТАЖЕННЯ ВСІХ ПРЕПАРАТІВ З RAW ФАЙЛІВ")
    print("=" * 60)

    all_drugs = []
    raw_files = list(RAW_DATA_PATH.glob("Rd2_*.csv"))

    print(f"Знайдено raw файлів: {len(raw_files)}")

    for raw_file in raw_files:
        try:
            # Читаємо тільки потрібні колонки (raw файли використовують ; як роздільник)
            df = pd.read_csv(raw_file, sep=';', usecols=RAW_COLUMNS_ORIGINAL, encoding='utf-8-sig')
            # Перейменовуємо колонки
            df = df.rename(columns=RAW_COLUMN_RENAME)
            all_drugs.append(df)
            print(f"  {raw_file.name}: {len(df)} рядків")
        except Exception as e:
            print(f"  ПОМИЛКА при читанні {raw_file.name}: {e}")

    if not all_drugs:
        raise ValueError("Не знайдено жодного raw файлу!")

    # Об'єднуємо та дедуплікуємо
    combined = pd.concat(all_drugs, ignore_index=True)
    unique_drugs = combined.drop_duplicates(subset=['DRUGS_ID']).copy()

    # Сортуємо по DRUGS_ID
    unique_drugs = unique_drugs.sort_values('DRUGS_ID').reset_index(drop=True)

    print(f"\nВсього унікальних препаратів: {len(unique_drugs)}")
    print(f"Унікальних INN груп: {unique_drugs['INN_ID'].nunique()}")

    return unique_drugs


def load_cross_market_data() -> Tuple[pd.DataFrame, List[int]]:
    """
    Завантажити всі cross_market CSV файли з Phase 1.

    Returns:
        Tuple: (об'єднаний DataFrame, список market_ids)
    """
    print("\n" + "=" * 60)
    print("ЗАВАНТАЖЕННЯ CROSS-MARKET ДАНИХ (Phase 1)")
    print("=" * 60)

    cross_market_files = list(CROSS_MARKET_PATH.glob("cross_market_*.csv"))

    if not cross_market_files:
        raise ValueError(f"Не знайдено cross_market файлів у {CROSS_MARKET_PATH}")

    print(f"Знайдено cross_market файлів: {len(cross_market_files)}")

    all_data = []
    market_ids = []

    for cm_file in sorted(cross_market_files):
        try:
            # Витягуємо market_id з назви файлу
            market_id = int(cm_file.stem.replace("cross_market_", ""))
            market_ids.append(market_id)

            df = pd.read_csv(cm_file)
            all_data.append(df)

            print(f"  {cm_file.name}: {len(df)} препаратів")
        except Exception as e:
            print(f"  ПОМИЛКА при читанні {cm_file.name}: {e}")

    combined = pd.concat(all_data, ignore_index=True)

    print(f"\nВсього записів: {len(combined)}")
    print(f"Унікальних препаратів: {combined['DRUGS_ID'].nunique()}")
    print(f"Ринків: {len(market_ids)}")

    return combined, sorted(market_ids)


# =============================================================================
# СТВОРЕННЯ ВИХІДНИХ ФАЙЛІВ
# =============================================================================

def create_all_drugs_list(raw_drugs: pd.DataFrame) -> pd.DataFrame:
    """
    Створити список всіх препаратів.

    Args:
        raw_drugs: DataFrame з усіма унікальними препаратами

    Returns:
        DataFrame для all_drugs_list.csv
    """
    print("\n" + "-" * 40)
    print("Створення all_drugs_list...")

    result = raw_drugs[['DRUGS_ID', 'DRUGS_NAME', 'INN_ID', 'INN_NAME']].copy()

    print(f"  Всього препаратів: {len(result)}")

    return result


def create_researched_drugs_list(
    cross_market_data: pd.DataFrame,
    market_ids: List[int]
) -> pd.DataFrame:
    """
    Створити список досліджених препаратів з coverage.

    Args:
        cross_market_data: Об'єднані cross_market дані
        market_ids: Список ID ринків

    Returns:
        DataFrame для researched_drugs_list.csv
    """
    print("\n" + "-" * 40)
    print("Створення researched_drugs_list...")

    total_markets = len(market_ids)

    # Агрегуємо по препаратах
    drug_stats = cross_market_data.groupby('DRUGS_ID').agg({
        'DRUGS_NAME': 'first',
        'INN_ID': 'first',
        'INN_NAME': 'first',
        'NFC1_ID': 'first',
        'CLIENT_ID': 'nunique'  # Кількість ринків
    }).reset_index()

    drug_stats.columns = ['DRUGS_ID', 'DRUGS_NAME', 'INN_ID', 'INN_NAME',
                          'NFC1_ID', 'MARKET_COUNT']

    # Розраховуємо coverage
    drug_stats['TOTAL_MARKETS'] = total_markets
    drug_stats['MARKET_COVERAGE'] = drug_stats['MARKET_COUNT'] / total_markets

    # Визначаємо кластер
    drug_stats['COVERAGE_CLUSTER'] = drug_stats['MARKET_COVERAGE'].apply(get_coverage_cluster)

    # Сортуємо по coverage (DESC) та DRUGS_ID
    drug_stats = drug_stats.sort_values(
        ['MARKET_COUNT', 'DRUGS_ID'],
        ascending=[False, True]
    ).reset_index(drop=True)

    print(f"  Всього досліджених препаратів: {len(drug_stats)}")
    print(f"  Кластери покриття:")
    for cluster in ['HIGH', 'MEDIUM', 'LOW', 'INSUFFICIENT']:
        count = (drug_stats['COVERAGE_CLUSTER'] == cluster).sum()
        print(f"    {cluster}: {count}")

    return drug_stats


def create_researched_drugs_coefficients(
    cross_market_data: pd.DataFrame,
    researched_drugs: pd.DataFrame,
    market_ids: List[int]
) -> pd.DataFrame:
    """
    Створити широкий формат даних з коефіцієнтами по ринках.

    Алгоритм "трикутного" впорядкування:
    1. Препарати відсортовані по MARKET_COUNT (DESC)
    2. Ринки відсортовані по кількості препаратів (DESC) для кращого "трикутника"

    Args:
        cross_market_data: Об'єднані cross_market дані
        researched_drugs: DataFrame з researched_drugs_list
        market_ids: Список ID ринків

    Returns:
        DataFrame для researched_drugs_coefficients.csv
    """
    print("\n" + "-" * 40)
    print("Створення researched_drugs_coefficients (трикутний формат)...")

    # Крок 1: Визначаємо порядок ринків по "заповненості"
    market_drug_counts = cross_market_data.groupby('CLIENT_ID')['DRUGS_ID'].nunique()
    sorted_markets = market_drug_counts.sort_values(ascending=False).index.tolist()

    print(f"  Порядок ринків по заповненості:")
    for i, mid in enumerate(sorted_markets[:5]):
        print(f"    {i+1}. Market {mid}: {market_drug_counts[mid]} препаратів")
    if len(sorted_markets) > 5:
        print(f"    ... ще {len(sorted_markets) - 5} ринків")

    # Крок 2: Створюємо базовий DataFrame з метаданими препаратів
    result = researched_drugs[['DRUGS_ID', 'DRUGS_NAME', 'INN_ID', 'INN_NAME',
                                'NFC1_ID', 'MARKET_COUNT']].copy()

    # Крок 3: Pivot для кожного ринку
    # Створюємо словник для швидкого доступу
    cross_market_dict = {}
    for _, row in cross_market_data.iterrows():
        key = (row['DRUGS_ID'], row['CLIENT_ID'])
        cross_market_dict[key] = {
            'SHARE_INTERNAL': row['SHARE_INTERNAL'],
            'INTERNAL_LIFT': row['INTERNAL_LIFT'],
            'EVENTS_COUNT': row['EVENTS_COUNT']
        }

    # Додаємо колонки для кожного ринку (в порядку заповненості)
    for market_id in sorted_markets:
        share_col = f'SHARE_INTERNAL_LOC_{market_id}'
        lift_col = f'INTERNAL_LIFT_LOC_{market_id}'
        events_col = f'EVENTS_COUNT_LOC_{market_id}'

        shares = []
        lifts = []
        events = []

        for drug_id in result['DRUGS_ID']:
            key = (drug_id, market_id)
            if key in cross_market_dict:
                data = cross_market_dict[key]
                shares.append(data['SHARE_INTERNAL'])
                lifts.append(data['INTERNAL_LIFT'])
                events.append(data['EVENTS_COUNT'])
            else:
                shares.append(np.nan)
                lifts.append(np.nan)
                events.append(np.nan)

        result[share_col] = shares
        result[lift_col] = lifts
        result[events_col] = events

    # Крок 4: Сортуємо препарати для "трикутного" вигляду
    # Вже відсортовані в researched_drugs по MARKET_COUNT DESC

    print(f"  Створено {len(result)} рядків")
    print(f"  Колонок: {len(result.columns)}")

    # Статистика заповненості
    data_cols = [c for c in result.columns if c.startswith('SHARE_INTERNAL_LOC_')]
    non_null_counts = result[data_cols].notna().sum(axis=1)
    print(f"  Середня заповненість: {non_null_counts.mean():.1f} ринків на препарат")
    print(f"  Макс заповненість: {non_null_counts.max()} ринків")
    print(f"  Мін заповненість: {non_null_counts.min()} ринків")

    return result


def create_coverage_analysis(
    all_drugs: pd.DataFrame,
    researched_drugs: pd.DataFrame,
    market_ids: List[int]
) -> pd.DataFrame:
    """
    Створити summary статистику покриття.

    Args:
        all_drugs: DataFrame з усіма препаратами
        researched_drugs: DataFrame з дослідженими препаратами
        market_ids: Список ID ринків

    Returns:
        DataFrame для coverage_analysis.csv
    """
    print("\n" + "-" * 40)
    print("Створення coverage_analysis...")

    metrics = []

    # Базова статистика
    metrics.append(('TOTAL_MARKETS', len(market_ids)))
    metrics.append(('TOTAL_DRUGS_RAW', len(all_drugs)))
    metrics.append(('TOTAL_DRUGS_RESEARCHED', len(researched_drugs)))

    # Coverage rate
    if len(all_drugs) > 0:
        coverage_rate = len(researched_drugs) / len(all_drugs)
    else:
        coverage_rate = 0.0
    metrics.append(('RAW_COVERAGE_RATE', round(coverage_rate, 4)))

    # По кластерах
    for cluster in ['HIGH', 'MEDIUM', 'LOW', 'INSUFFICIENT']:
        count = (researched_drugs['COVERAGE_CLUSTER'] == cluster).sum()
        metrics.append((f'DRUGS_{cluster}_COVERAGE', count))

    # Додаткова статистика
    if len(researched_drugs) > 0:
        avg_market_count = researched_drugs['MARKET_COUNT'].mean()
        metrics.append(('AVG_MARKETS_PER_DRUG', round(avg_market_count, 2)))

        avg_coverage = researched_drugs['MARKET_COVERAGE'].mean()
        metrics.append(('AVG_MARKET_COVERAGE', round(avg_coverage, 4)))

    result = pd.DataFrame(metrics, columns=['METRIC', 'VALUE'])

    print(f"  Метрик: {len(result)}")
    for _, row in result.iterrows():
        print(f"    {row['METRIC']}: {row['VALUE']}")

    return result


# =============================================================================
# ВАЛІДАЦІЯ
# =============================================================================

def validate_results(
    all_drugs: pd.DataFrame,
    researched_drugs: pd.DataFrame,
    coefficients: pd.DataFrame,
    coverage_analysis: pd.DataFrame,
    market_ids: List[int]
) -> Tuple[bool, List[str]]:
    """
    Валідація всіх результатів.

    Returns:
        Tuple: (всі тести пройшли, список повідомлень)
    """
    print("\n" + "=" * 60)
    print("ВАЛІДАЦІЯ РЕЗУЛЬТАТІВ")
    print("=" * 60)

    messages = []
    all_passed = True

    # 1. TOTAL_MARKETS
    total_markets_reported = coverage_analysis[
        coverage_analysis['METRIC'] == 'TOTAL_MARKETS'
    ]['VALUE'].values[0]

    if total_markets_reported == len(market_ids):
        messages.append(f"[OK] TOTAL_MARKETS: {total_markets_reported} = {len(market_ids)} (cross_market files)")
    else:
        messages.append(f"[FAIL] TOTAL_MARKETS mismatch: {total_markets_reported} != {len(market_ids)}")
        all_passed = False

    # 2. DRUGS_COVERAGE сума
    cluster_sum = 0
    for cluster in ['HIGH', 'MEDIUM', 'LOW', 'INSUFFICIENT']:
        metric_name = f'DRUGS_{cluster}_COVERAGE'
        val = coverage_analysis[coverage_analysis['METRIC'] == metric_name]['VALUE'].values
        if len(val) > 0:
            cluster_sum += val[0]

    total_researched = len(researched_drugs)
    if cluster_sum == total_researched:
        messages.append(f"[OK] COVERAGE_CLUSTERS sum: {cluster_sum} = {total_researched} (total researched)")
    else:
        messages.append(f"[FAIL] COVERAGE_CLUSTERS sum mismatch: {cluster_sum} != {total_researched}")
        all_passed = False

    # 3. SHARE_INTERNAL в діапазоні [0, 1]
    share_cols = [c for c in coefficients.columns if c.startswith('SHARE_INTERNAL_LOC_')]
    all_values = coefficients[share_cols].values.flatten()
    valid_values = all_values[~np.isnan(all_values)]

    out_of_range = ((valid_values < 0) | (valid_values > 1)).sum()
    if out_of_range == 0:
        messages.append(f"[OK] SHARE_INTERNAL range: all {len(valid_values)} values in [0, 1]")
    else:
        messages.append(f"[FAIL] SHARE_INTERNAL range: {out_of_range} values out of [0, 1]")
        all_passed = False

    # 4. Трикутна структура (перший препарат має максимальне заповнення)
    non_null_counts = coefficients[share_cols].notna().sum(axis=1)
    if len(non_null_counts) > 1:
        if non_null_counts.iloc[0] >= non_null_counts.iloc[1]:
            messages.append(f"[OK] TRIANGLE_STRUCTURE: first drug has max coverage ({non_null_counts.iloc[0]})")
        else:
            messages.append(f"[WARN] TRIANGLE_STRUCTURE: first drug ({non_null_counts.iloc[0]}) < second ({non_null_counts.iloc[1]})")

    # 5. Ринки в coefficients відповідають Phase 1
    coef_markets = set()
    for col in share_cols:
        market_id = int(col.replace('SHARE_INTERNAL_LOC_', ''))
        coef_markets.add(market_id)

    if coef_markets == set(market_ids):
        messages.append(f"[OK] MARKET_MATCH: all {len(market_ids)} markets present in coefficients")
    else:
        missing = set(market_ids) - coef_markets
        extra = coef_markets - set(market_ids)
        messages.append(f"[FAIL] MARKET_MATCH: missing={missing}, extra={extra}")
        all_passed = False

    # 6. MARKET_COUNT відповідає реальній кількості значень
    for idx, row in coefficients.head(5).iterrows():
        actual_count = coefficients.loc[idx, share_cols].notna().sum()
        reported_count = row['MARKET_COUNT']
        if actual_count == reported_count:
            messages.append(f"[OK] MARKET_COUNT for drug {row['DRUGS_ID']}: {actual_count}")
        else:
            messages.append(f"[FAIL] MARKET_COUNT for drug {row['DRUGS_ID']}: actual={actual_count}, reported={reported_count}")
            all_passed = False

    # Виводимо результати
    for msg in messages:
        print(f"  {msg}")

    status = "PASSED" if all_passed else "FAILED"
    print(f"\nЗАГАЛЬНИЙ СТАТУС: {status}")

    return all_passed, messages


def create_validation_report(
    all_passed: bool,
    messages: List[str],
    output_path: Path
) -> None:
    """
    Створити текстовий файл валідації.

    Args:
        all_passed: Чи всі тести пройшли
        messages: Список повідомлень валідації
        output_path: Шлях для збереження
    """
    report_path = output_path / "validation_report.txt"

    with open(report_path, 'w', encoding='utf-8') as f:
        f.write("=" * 70 + "\n")
        f.write("VALIDATION REPORT - Phase 2 Step 1: Data Preparation\n")
        f.write(f"Generated: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}\n")
        f.write("=" * 70 + "\n\n")

        status = "PASSED" if all_passed else "FAILED"
        f.write(f"OVERALL STATUS: {status}\n\n")

        f.write("-" * 40 + "\n")
        f.write("VALIDATION CHECKS:\n")
        f.write("-" * 40 + "\n")

        for msg in messages:
            f.write(f"{msg}\n")

        f.write("\n" + "=" * 70 + "\n")

    print(f"\nValidation report saved: {report_path}")


# =============================================================================
# ЕКСПОРТ
# =============================================================================

def export_to_csv(
    all_drugs: pd.DataFrame,
    researched_drugs: pd.DataFrame,
    coefficients: pd.DataFrame,
    coverage_analysis: pd.DataFrame,
    output_path: Path
) -> None:
    """
    Експортувати всі дані в CSV.
    """
    print("\n" + "-" * 40)
    print("Експорт в CSV...")

    # Створюємо папку якщо не існує
    output_path.mkdir(parents=True, exist_ok=True)

    # all_drugs_list
    all_drugs.to_csv(output_path / "all_drugs_list.csv", index=False)
    print(f"  all_drugs_list.csv: {len(all_drugs)} рядків")

    # researched_drugs_list
    researched_drugs.to_csv(output_path / "researched_drugs_list.csv", index=False)
    print(f"  researched_drugs_list.csv: {len(researched_drugs)} рядків")

    # researched_drugs_coefficients
    coefficients.to_csv(output_path / "researched_drugs_coefficients.csv", index=False)
    print(f"  researched_drugs_coefficients.csv: {len(coefficients)} рядків, {len(coefficients.columns)} колонок")

    # coverage_analysis
    coverage_analysis.to_csv(output_path / "coverage_analysis.csv", index=False)
    print(f"  coverage_analysis.csv: {len(coverage_analysis)} рядків")


def export_to_xlsx(
    researched_drugs: pd.DataFrame,
    coefficients: pd.DataFrame,
    coverage_analysis: pd.DataFrame,
    output_path: Path
) -> None:
    """
    Експортувати дані в Excel для бізнес-презентації.
    """
    print("\n" + "-" * 40)
    print("Експорт в XLSX для бізнесу...")

    # Створюємо папку якщо не існує
    output_path.mkdir(parents=True, exist_ok=True)

    try:
        # researched_drugs_list.xlsx
        researched_drugs.to_excel(
            output_path / "researched_drugs_list.xlsx",
            index=False,
            sheet_name="Researched Drugs"
        )
        print(f"  researched_drugs_list.xlsx: {len(researched_drugs)} рядків")

        # coverage_analysis.xlsx
        coverage_analysis.to_excel(
            output_path / "coverage_analysis.xlsx",
            index=False,
            sheet_name="Coverage Analysis"
        )
        print(f"  coverage_analysis.xlsx: {len(coverage_analysis)} рядків")

        # researched_drugs_coefficients.xlsx
        # Для бізнес-звіту прибираємо технічні колонки INTERNAL_LIFT_LOC_ та EVENTS_COUNT_LOC_
        # (замовник бачить тільки SHARE_INTERNAL_LOC_ по ринках)
        # Ці колонки залишаються в CSV для подальших розрахунків (зважений коефіцієнт)
        business_cols = [
            c for c in coefficients.columns
            if not c.startswith('INTERNAL_LIFT_LOC_') and not c.startswith('EVENTS_COUNT_LOC_')
        ]
        coefficients_business = coefficients[business_cols].copy()

        # Зберігаємо xlsx, потім форматуємо SHARE_INTERNAL_LOC_ як відсотки через openpyxl
        xlsx_path = output_path / "researched_drugs_coefficients.xlsx"
        coefficients_business.to_excel(xlsx_path, index=False, sheet_name="Coefficients")

        # Форматування: значення 0.3674 відображається як 36.74%
        from openpyxl import load_workbook
        wb = load_workbook(xlsx_path)
        ws = wb.active
        share_col_indices = [
            i + 1 for i, c in enumerate(business_cols)
            if c.startswith('SHARE_INTERNAL_LOC_')
        ]
        for col_idx in share_col_indices:
            for row in range(2, ws.max_row + 1):  # рядок 1 = заголовок
                cell = ws.cell(row=row, column=col_idx)
                if cell.value is not None:
                    cell.number_format = '0.00%'
        wb.save(xlsx_path)

        print(f"  researched_drugs_coefficients.xlsx: {len(coefficients_business)} рядків, {len(business_cols)} колонок (SHARE як %, без INTERNAL_LIFT_LOC_, EVENTS_COUNT_LOC_)")

    except Exception as e:
        print(f"  ПОМИЛКА при експорті XLSX: {e}")
        print("  (Можливо потрібно встановити openpyxl: pip install openpyxl)")


# =============================================================================
# MAIN
# =============================================================================

def main():
    """Головна функція."""
    print("=" * 70)
    print("PHASE 2, STEP 1: DATA PREPARATION")
    print(f"Started: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}")
    print("=" * 70)

    # 1. Завантаження даних
    all_drugs = load_all_drugs_from_raw()
    cross_market_data, market_ids = load_cross_market_data()

    # 2. Створення вихідних файлів
    all_drugs_list = create_all_drugs_list(all_drugs)
    researched_drugs = create_researched_drugs_list(cross_market_data, market_ids)
    coefficients = create_researched_drugs_coefficients(
        cross_market_data, researched_drugs, market_ids
    )
    coverage_analysis = create_coverage_analysis(all_drugs, researched_drugs, market_ids)

    # 3. Валідація
    all_passed, messages = validate_results(
        all_drugs_list, researched_drugs, coefficients,
        coverage_analysis, market_ids
    )

    # 4. Експорт CSV
    export_to_csv(
        all_drugs_list, researched_drugs, coefficients,
        coverage_analysis, OUTPUT_BASE_PATH
    )

    # 5. Експорт XLSX
    export_to_xlsx(
        researched_drugs, coefficients, coverage_analysis,
        OUTPUT_BUSINESS_PATH
    )

    # 6. Validation report
    create_validation_report(all_passed, messages, OUTPUT_BASE_PATH)

    # Summary
    print("\n" + "=" * 70)
    print("SUMMARY")
    print("=" * 70)
    print(f"  All drugs (raw): {len(all_drugs_list)}")
    print(f"  Researched drugs: {len(researched_drugs)}")
    print(f"  Markets: {len(market_ids)}")
    print(f"  Validation: {'PASSED' if all_passed else 'FAILED'}")
    print(f"\nOutput folder: {OUTPUT_BASE_PATH}")
    print(f"Business reports: {OUTPUT_BUSINESS_PATH}")
    print(f"\nFinished: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}")
    print("=" * 70)


if __name__ == "__main__":
    main()
