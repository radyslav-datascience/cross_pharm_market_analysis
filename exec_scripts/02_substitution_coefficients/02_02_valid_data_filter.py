"""
02_02_valid_data_filter.py - Фільтрація валідних препаратів для бізнес-висновків

Phase 2, Step 2.2 part 2: Valid Data Filter (Scenario A — strict)

Критерії фільтрації:
    - COVERAGE_CLUSTER ∈ {HIGH, MEDIUM}
    - RELIABILITY ∈ {HIGH, MEDIUM}

Вхідні дані:
- results/substitution_research/02_statistics_and_filter/02_01_statistical_analysis/drug_statistics.csv
- results/substitution_research/01_preparation/all_drugs_list.csv
- data/raw/Rd2_*.csv (для підрахунку загальної кількості DRUGS_ID)

Вихідні дані:
- results/substitution_research/02_statistics_and_filter/02_02_valid_data_filter/valid_drugs.csv
- results/substitution_research/02_statistics_and_filter/02_02_valid_data_filter/rejected_drugs.csv
- results/substitution_research/02_statistics_and_filter/02_02_valid_data_filter/filter_summary.csv
- results/substitution_research/02_statistics_and_filter/02_02_valid_data_filter/validation_report.txt
- results/substitution_research/02_statistics_and_filter/02_02_valid_data_filter/filter_business_reports/valid_drugs.xlsx
- results/substitution_research/02_statistics_and_filter/02_02_valid_data_filter/filter_business_reports/rejected_drugs.xlsx
- results/substitution_research/02_statistics_and_filter/02_02_valid_data_filter/filter_business_reports/cross_table_distribution.xlsx
- results/substitution_research/02_statistics_and_filter/02_02_valid_data_filter/filter_business_reports/pipeline_funnel.xlsx

Використання:
    python exec_scripts/02_substitution_coefficients/02_02_valid_data_filter.py
"""

import pandas as pd
import numpy as np
import sys
import glob
from pathlib import Path
from datetime import datetime
from typing import List, Dict, Tuple

# Додаємо шлях до project_core
SCRIPT_PATH = Path(__file__).resolve()
PROJECT_ROOT = SCRIPT_PATH.parent.parent.parent
sys.path.insert(0, str(PROJECT_ROOT))

from project_core.data_config.paths_config import RESULTS_PATH, RAW_DATA_PATH
from project_core.sub_coef_config.coverage_thresholds import (
    COVERAGE_HIGH, COVERAGE_MEDIUM,
    get_coverage_cluster, get_cluster_description
)
from project_core.sub_coef_config.reliability_thresholds import (
    RELIABILITY_HIGH, RELIABILITY_MEDIUM,
    get_reliability_class, get_reliability_description
)


# =============================================================================
# КОНСТАНТИ
# =============================================================================

# Вхідні шляхи
STATS_INPUT_PATH = (
    RESULTS_PATH / "substitution_research" / "02_statistics_and_filter"
    / "02_01_statistical_analysis" / "drug_statistics.csv"
)
ALL_DRUGS_PATH = (
    RESULTS_PATH / "substitution_research" / "01_preparation" / "all_drugs_list.csv"
)

# Вихідні шляхи
OUTPUT_BASE_PATH = (
    RESULTS_PATH / "substitution_research" / "02_statistics_and_filter"
    / "02_02_valid_data_filter"
)
OUTPUT_BUSINESS_PATH = OUTPUT_BASE_PATH / "filter_business_reports"

# Фільтр Scenario A (strict)
VALID_COVERAGE_CLUSTERS = {'HIGH', 'MEDIUM'}
VALID_RELIABILITY_CLASSES = {'HIGH', 'MEDIUM'}

# Порядок категорій для cross-table
COVERAGE_ORDER = ['HIGH', 'MEDIUM', 'LOW', 'INSUFFICIENT']
RELIABILITY_ORDER = ['HIGH', 'MEDIUM', 'LOW', 'SINGLE_MARKET']


# =============================================================================
# ЗАВАНТАЖЕННЯ ДАНИХ
# =============================================================================

def load_drug_statistics() -> pd.DataFrame:
    """Завантажити drug_statistics.csv з попереднього кроку."""
    print("\nЗавантаження drug_statistics.csv...")

    if not STATS_INPUT_PATH.exists():
        raise FileNotFoundError(
            f"Файл не знайдено: {STATS_INPUT_PATH}\n"
            f"Спочатку виконайте: python exec_scripts/02_substitution_coefficients/"
            f"02_01_statistical_analysis.py"
        )

    df = pd.read_csv(STATS_INPUT_PATH)
    print(f"  Завантажено: {len(df)} препаратів, {len(df.columns)} колонок")
    return df


def load_all_drugs_list() -> pd.DataFrame:
    """Завантажити all_drugs_list.csv."""
    print("Завантаження all_drugs_list.csv...")

    if not ALL_DRUGS_PATH.exists():
        raise FileNotFoundError(
            f"Файл не знайдено: {ALL_DRUGS_PATH}\n"
            f"Спочатку виконайте Phase 2 Step 1."
        )

    df = pd.read_csv(ALL_DRUGS_PATH)
    print(f"  Завантажено: {len(df)} препаратів (повна популяція)")
    return df


def count_raw_unique_drugs() -> int:
    """Підрахувати унікальні DRUGS_ID в raw файлах data/raw/Rd2_*.csv."""
    print("Підрахунок унікальних DRUGS_ID в raw файлах...")

    raw_pattern = str(RAW_DATA_PATH / "Rd2_*.csv")
    raw_files = sorted(glob.glob(raw_pattern))

    if not raw_files:
        raise FileNotFoundError(f"Raw файли не знайдено: {raw_pattern}")

    all_drug_ids = set()
    for f in raw_files:
        df = pd.read_csv(f, sep=';', usecols=['DRUGS_ID'], encoding='utf-8-sig')
        all_drug_ids.update(df['DRUGS_ID'].unique())

    print(f"  Raw файлів: {len(raw_files)}")
    print(f"  Унікальних DRUGS_ID: {len(all_drug_ids)}")
    return len(all_drug_ids)


# =============================================================================
# ФІЛЬТРАЦІЯ (SCENARIO A — STRICT)
# =============================================================================

def apply_filter(
    drug_stats: pd.DataFrame,
    valid_coverage: set = VALID_COVERAGE_CLUSTERS,
    valid_reliability: set = VALID_RELIABILITY_CLASSES
) -> Tuple[pd.DataFrame, pd.DataFrame]:
    """
    Фільтрація препаратів за Scenario A (strict).

    Критерії:
        COVERAGE_CLUSTER ∈ {HIGH, MEDIUM} AND RELIABILITY ∈ {HIGH, MEDIUM}

    Args:
        drug_stats: DataFrame з drug_statistics.csv
        valid_coverage: Допустимі кластери покриття
        valid_reliability: Допустимі класи надійності

    Returns:
        Tuple[valid_drugs, rejected_drugs]: Два DataFrame
    """
    print("\n" + "-" * 40)
    print("Фільтрація Scenario A (strict)...")
    print(f"  Критерії:")
    print(f"    COVERAGE_CLUSTER ∈ {sorted(valid_coverage)}")
    print(f"    RELIABILITY ∈ {sorted(valid_reliability)}")

    mask_coverage = drug_stats['COVERAGE_CLUSTER'].isin(valid_coverage)
    mask_reliability = drug_stats['RELIABILITY'].isin(valid_reliability)
    mask_valid = mask_coverage & mask_reliability

    valid_drugs = drug_stats[mask_valid].copy()
    rejected_drugs = drug_stats[~mask_valid].copy()

    # Додаємо колонку FILTER_STATUS
    valid_drugs.insert(0, 'FILTER_STATUS', 'VALID')
    rejected_drugs.insert(0, 'FILTER_STATUS', 'REJECTED')

    # Додаємо причину відхилення
    reject_reasons = []
    for _, row in rejected_drugs.iterrows():
        reasons = []
        if row['COVERAGE_CLUSTER'] not in valid_coverage:
            reasons.append(f"COVERAGE={row['COVERAGE_CLUSTER']}")
        if row['RELIABILITY'] not in valid_reliability:
            reasons.append(f"RELIABILITY={row['RELIABILITY']}")
        reject_reasons.append('; '.join(reasons))

    rejected_drugs['REJECT_REASON'] = reject_reasons

    # Перевпорядкування колонок логічними блоками
    column_order = [
        # ID
        'FILTER_STATUS', 'DRUGS_ID', 'DRUGS_NAME', 'INN_ID', 'INN_NAME', 'NFC1_ID',
        # Coverage block
        'COVERAGE_CLUSTER', 'MARKET_COVERAGE', 'MARKET_COUNT_TOTAL', 'MARKET_COUNT_CLEAN', 'OUTLIERS_COUNT',
        # Reliability block
        'RELIABILITY', 'VARIATION_COEFFICIENT', 'STD_SHARE_INTERNAL',
        # Центральні метрики
        'MEDIAN_SHARE_INTERNAL', 'MEAN_SHARE_INTERNAL', 'WEIGHTED_MEAN_SHARE',
        # CI & розподіл
        'CI_95_LOWER', 'CI_95_UPPER', 'MIN_SHARE_INTERNAL', 'Q1_SHARE_INTERNAL',
        'Q3_SHARE_INTERNAL', 'MAX_SHARE_INTERNAL', 'IQR_SHARE_INTERNAL',
        # Обсяг
        'TOTAL_EVENTS', 'TOTAL_INTERNAL_LIFT',
    ]

    # Для rejected — додаємо REJECT_REASON в кінець
    valid_cols = [c for c in column_order if c in valid_drugs.columns]
    rejected_cols = [c for c in column_order if c in rejected_drugs.columns] + ['REJECT_REASON']

    valid_drugs = valid_drugs[valid_cols]
    rejected_drugs = rejected_drugs[rejected_cols]

    print(f"\n  Результат:")
    print(f"    VALID:    {len(valid_drugs):>4} препаратів ({len(valid_drugs)/len(drug_stats):.1%})")
    print(f"    REJECTED: {len(rejected_drugs):>4} препаратів ({len(rejected_drugs)/len(drug_stats):.1%})")
    print(f"    TOTAL:    {len(drug_stats):>4}")

    return valid_drugs, rejected_drugs


# =============================================================================
# CROSS-TABLE DISTRIBUTION
# =============================================================================

def create_cross_table(drug_stats: pd.DataFrame) -> Dict[str, pd.DataFrame]:
    """
    Створити крос-таблицю COVERAGE_CLUSTER × RELIABILITY.

    Повертає словник з трьома DataFrame:
        - 'counts': абсолютні значення
        - 'ratios': частки від загальної кількості
        - 'combined': текст "count (ratio%)"

    Args:
        drug_stats: DataFrame з drug_statistics.csv

    Returns:
        Dict з ключами 'counts', 'ratios', 'combined'
    """
    print("\n" + "-" * 40)
    print("Cross-table: COVERAGE_CLUSTER × RELIABILITY...")

    total = len(drug_stats)

    # Абсолютні значення
    ct_counts = pd.crosstab(
        drug_stats['COVERAGE_CLUSTER'],
        drug_stats['RELIABILITY'],
        margins=True,
        margins_name='TOTAL'
    )

    # Впорядковуємо рядки та стовпці
    row_order = [c for c in COVERAGE_ORDER if c in ct_counts.index] + ['TOTAL']
    col_order = [c for c in RELIABILITY_ORDER if c in ct_counts.columns] + ['TOTAL']
    ct_counts = ct_counts.reindex(index=row_order, columns=col_order, fill_value=0)

    # Підпис осей: index.name = рядки (COVERAGE), columns.name = стовпці (RELIABILITY)
    ct_counts.index.name = 'COVERAGE ↓ \\ RELIABILITY →'
    ct_counts.columns.name = None

    # Частки (ratios) — відносно загальної кількості досліджених препаратів
    ct_ratios = ct_counts / total
    ct_ratios.index.name = ct_counts.index.name

    # Комбінований текст
    ct_combined = ct_counts.copy().astype(str)
    ct_combined.index.name = ct_counts.index.name
    for row in ct_combined.index:
        for col in ct_combined.columns:
            count_val = ct_counts.loc[row, col]
            ratio_val = ct_ratios.loc[row, col]
            ct_combined.loc[row, col] = f"{count_val} ({ratio_val:.1%})"

    print(f"  Розмір таблиці: {ct_counts.shape[0]} рядків × {ct_counts.shape[1]} колонок")
    print(f"\n  Підсумок:")
    for cov in COVERAGE_ORDER:
        if cov in ct_counts.index:
            row_total = ct_counts.loc[cov, 'TOTAL']
            print(f"    {cov:>15}: {row_total:>4} ({row_total/total:.1%})")

    return {
        'counts': ct_counts,
        'ratios': ct_ratios,
        'combined': ct_combined
    }


# =============================================================================
# PIPELINE FUNNEL
# =============================================================================

def create_pipeline_funnel(
    raw_drugs_count: int,
    all_drugs_count: int,
    researched_count: int,
    valid_count: int,
    total_markets: int
) -> pd.DataFrame:
    """
    Створити воронку pipeline з інтерполяцією.

    Воронка:
        1. Total drugs in raw data
        2. Researched drugs (пройшли Phase 1)
        3. Valid drugs (пройшли фільтр Scenario A)
        4. Інтерполяція: скільки додаткових ринків потрібно

    Args:
        raw_drugs_count: Унікальних DRUGS_ID в raw файлах
        all_drugs_count: Записів в all_drugs_list.csv
        researched_count: Кількість досліджених (drug_statistics.csv)
        valid_count: Кількість після фільтру
        total_markets: Поточна кількість ринків

    Returns:
        pd.DataFrame: Таблиця воронки
    """
    print("\n" + "-" * 40)
    print("Pipeline funnel + інтерполяція...")

    # Основна воронка
    funnel_data = []

    # Рядок 1: Raw drugs
    funnel_data.append({
        'STAGE': '1. Усього препаратів у raw даних',
        'COUNT': raw_drugs_count,
        'RATIO_VS_RAW': 1.0,
        'RATIO_VS_RESEARCHED': None,
        'NOTE': f'Унікальних DRUGS_ID у {total_markets} ринках'
    })

    # Рядок 2: All drugs (повинно = raw)
    # Можуть відрізнятись, тому показуємо обидва
    if all_drugs_count != raw_drugs_count:
        funnel_data.append({
            'STAGE': '1b. all_drugs_list.csv (верифікація)',
            'COUNT': all_drugs_count,
            'RATIO_VS_RAW': all_drugs_count / raw_drugs_count if raw_drugs_count > 0 else None,
            'RATIO_VS_RESEARCHED': None,
            'NOTE': 'Має збігатися з raw (контроль цілісності)'
        })

    # Рядок 3: Researched
    funnel_data.append({
        'STAGE': '2. Досліджені препарати (Phase 1)',
        'COUNT': researched_count,
        'RATIO_VS_RAW': researched_count / raw_drugs_count if raw_drugs_count > 0 else None,
        'RATIO_VS_RESEARCHED': 1.0,
        'NOTE': 'Мали хоча б 1 stock-out event на хоча б 1 ринку'
    })

    # Рядок 4: Valid (passed filter)
    funnel_data.append({
        'STAGE': '3. Пройшли фільтр (Scenario A)',
        'COUNT': valid_count,
        'RATIO_VS_RAW': valid_count / raw_drugs_count if raw_drugs_count > 0 else None,
        'RATIO_VS_RESEARCHED': valid_count / researched_count if researched_count > 0 else None,
        'NOTE': 'COVERAGE ∈ {HIGH, MEDIUM} AND RELIABILITY ∈ {HIGH, MEDIUM}'
    })

    # Рядок 5: Rejected
    rejected_count = researched_count - valid_count
    funnel_data.append({
        'STAGE': '4. Відхилені фільтром',
        'COUNT': rejected_count,
        'RATIO_VS_RAW': rejected_count / raw_drugs_count if raw_drugs_count > 0 else None,
        'RATIO_VS_RESEARCHED': rejected_count / researched_count if researched_count > 0 else None,
        'NOTE': 'COVERAGE ∈ {LOW, INSUFFICIENT} або RELIABILITY ∈ {LOW, SINGLE_MARKET}'
    })

    # Рядок 6: Not researched (не потрапили в дослідження)
    not_researched = raw_drugs_count - researched_count
    funnel_data.append({
        'STAGE': '5. Не потрапили в дослідження',
        'COUNT': not_researched,
        'RATIO_VS_RAW': not_researched / raw_drugs_count if raw_drugs_count > 0 else None,
        'RATIO_VS_RESEARCHED': None,
        'NOTE': 'Немає stock-out events на жодному з ринків'
    })

    funnel_df = pd.DataFrame(funnel_data)

    # --- ІНТЕРПОЛЯЦІЯ ---
    # Логарифмічна модель: rate(N) = rate_current * ln(N) / ln(N_current)
    # Для rate = 1.0: N_target = N_current^(1/rate_current)

    research_rate = researched_count / raw_drugs_count if raw_drugs_count > 0 else 0
    filter_rate = valid_count / raw_drugs_count if raw_drugs_count > 0 else 0

    # Оцінка 1: Research coverage (507/652) — щоб усі потрапили в дослідження
    if research_rate > 0 and total_markets > 1:
        n_research = total_markets ** (1.0 / research_rate)
        additional_research = max(0, int(np.ceil(n_research - total_markets)))
    else:
        n_research = None
        additional_research = None

    # Оцінка 2: Filter pass rate (242/652) — щоб усі пройшли фільтр
    if filter_rate > 0 and total_markets > 1:
        n_filter = total_markets ** (1.0 / filter_rate)
        additional_filter = max(0, int(np.ceil(n_filter - total_markets)))
    else:
        n_filter = None
        additional_filter = None

    interpolation_data = [
        # --- Вхідні дані ---
        {
            'PARAMETER': 'Поточна кількість ринків',
            'VALUE': total_markets,
            'NOTE': ''
        },
        {
            'PARAMETER': 'Усього препаратів (raw)',
            'VALUE': raw_drugs_count,
            'NOTE': 'Унікальних DRUGS_ID у всіх raw файлах'
        },
        {
            'PARAMETER': 'Досліджених препаратів',
            'VALUE': researched_count,
            'NOTE': f'Research rate = {research_rate:.1%}'
        },
        {
            'PARAMETER': 'Пройшли фільтр (Scenario A)',
            'VALUE': valid_count,
            'NOTE': f'Filter rate = {filter_rate:.1%}'
        },
        {
            'PARAMETER': '',
            'VALUE': '',
            'NOTE': ''
        },
        # --- Оцінка 1: Research coverage ---
        {
            'PARAMETER': '--- ОЦІНКА 1: RESEARCH COVERAGE ---',
            'VALUE': '',
            'NOTE': 'Скільки ринків щоб усі 652 потрапили в дослідження'
        },
        {
            'PARAMETER': 'Research rate',
            'VALUE': round(research_rate, 4),
            'NOTE': f'{researched_count} / {raw_drugs_count}'
        },
        {
            'PARAMETER': 'Модель',
            'VALUE': 'Логарифмічна',
            'NOTE': f'N_target = {total_markets}^(1/{research_rate:.4f})'
        },
        {
            'PARAMETER': 'Оцінка ринків для 100% дослідження',
            'VALUE': int(np.ceil(n_research)) if n_research else 'N/A',
            'NOTE': f'Усі {raw_drugs_count} препаратів матимуть хоча б 1 stock-out event'
        },
        {
            'PARAMETER': 'Додаткових ринків потрібно',
            'VALUE': additional_research if additional_research is not None else 'N/A',
            'NOTE': ''
        },
        {
            'PARAMETER': '',
            'VALUE': '',
            'NOTE': ''
        },
        # --- Оцінка 2: Filter pass rate ---
        {
            'PARAMETER': '--- ОЦІНКА 2: FILTER PASS RATE ---',
            'VALUE': '',
            'NOTE': 'Скільки ринків щоб усі 652 пройшли фільтр Scenario A'
        },
        {
            'PARAMETER': 'Filter rate',
            'VALUE': round(filter_rate, 4),
            'NOTE': f'{valid_count} / {raw_drugs_count}'
        },
        {
            'PARAMETER': 'Модель',
            'VALUE': 'Логарифмічна',
            'NOTE': f'N_target = {total_markets}^(1/{filter_rate:.4f})'
        },
        {
            'PARAMETER': 'Оцінка ринків для 100% проходження фільтру',
            'VALUE': int(np.ceil(n_filter)) if n_filter else 'N/A',
            'NOTE': 'ТЕОРЕТИЧНА ОЦІНКА — див. обмеження нижче'
        },
        {
            'PARAMETER': 'Додаткових ринків потрібно',
            'VALUE': additional_filter if additional_filter is not None else 'N/A',
            'NOTE': ''
        },
        {
            'PARAMETER': '',
            'VALUE': '',
            'NOTE': ''
        },
        # --- Обмеження ---
        {
            'PARAMETER': '--- ОБМЕЖЕННЯ МОДЕЛІ ---',
            'VALUE': '',
            'NOTE': ''
        },
        {
            'PARAMETER': 'Оцінка 1 (research)',
            'VALUE': 'Реалістична',
            'NOTE': 'Більше ринків → більше stock-out events → більше препаратів досліджено'
        },
        {
            'PARAMETER': 'Оцінка 2 (filter)',
            'VALUE': 'Нереалістична',
            'NOTE': (
                'RELIABILITY залежить від природної варіативності субституції препарату, '
                'а не лише від кількості ринків. Препарати з нестабільною субституцією '
                'не пройдуть фільтр незалежно від кількості даних.'
            )
        },
        {
            'PARAMETER': 'Висновок',
            'VALUE': '',
            'NOTE': (
                'Оцінка 1 показує мінімум ринків для повного дослідження. '
                'Оцінка 2 демонструє що 100% проходження фільтру недосяжне лише через додавання ринків.'
            )
        },
    ]

    interpolation_df = pd.DataFrame(interpolation_data)

    print(f"  Воронка:")
    for _, row in funnel_df.iterrows():
        ratio_str = f"({row['RATIO_VS_RAW']:.1%})" if pd.notna(row['RATIO_VS_RAW']) else ""
        print(f"    {row['STAGE']}: {row['COUNT']} {ratio_str}")

    print(f"\n  Інтерполяція:")
    print(f"    Оцінка 1 (research): ~{int(np.ceil(n_research))} ринків (+{additional_research})")
    if n_filter:
        print(f"    Оцінка 2 (filter):   ~{int(np.ceil(n_filter))} ринків (+{additional_filter}) — нереалістична")

    return funnel_df, interpolation_df


# =============================================================================
# FILTER SUMMARY
# =============================================================================

def create_filter_summary(
    drug_stats: pd.DataFrame,
    valid_drugs: pd.DataFrame,
    rejected_drugs: pd.DataFrame
) -> pd.DataFrame:
    """Створити summary метрики фільтрації."""

    total = len(drug_stats)
    valid_n = len(valid_drugs)
    rejected_n = len(rejected_drugs)

    summary_rows = [
        {'METRIC': 'TOTAL_DRUGS', 'VALUE': total, 'DESCRIPTION': 'Всього препаратів на вході'},
        {'METRIC': 'VALID_DRUGS', 'VALUE': valid_n, 'DESCRIPTION': 'Пройшли фільтр'},
        {'METRIC': 'REJECTED_DRUGS', 'VALUE': rejected_n, 'DESCRIPTION': 'Відхилені'},
        {'METRIC': 'VALID_RATIO', 'VALUE': round(valid_n / total, 4), 'DESCRIPTION': 'Частка валідних'},
        {'METRIC': 'FILTER_TYPE', 'VALUE': 'SCENARIO_A_STRICT', 'DESCRIPTION': 'Тип фільтру'},
        {'METRIC': 'COVERAGE_CRITERIA', 'VALUE': 'HIGH, MEDIUM', 'DESCRIPTION': 'Допустимі COVERAGE_CLUSTER'},
        {'METRIC': 'RELIABILITY_CRITERIA', 'VALUE': 'HIGH, MEDIUM', 'DESCRIPTION': 'Допустимі RELIABILITY'},
        {'METRIC': 'COVERAGE_HIGH_THRESHOLD', 'VALUE': COVERAGE_HIGH, 'DESCRIPTION': f'≥{COVERAGE_HIGH:.0%} ринків'},
        {'METRIC': 'COVERAGE_MEDIUM_THRESHOLD', 'VALUE': COVERAGE_MEDIUM, 'DESCRIPTION': f'≥{COVERAGE_MEDIUM:.0%} ринків'},
        {'METRIC': 'RELIABILITY_HIGH_THRESHOLD', 'VALUE': RELIABILITY_HIGH, 'DESCRIPTION': f'VARIATION_COEFFICIENT < {RELIABILITY_HIGH}'},
        {'METRIC': 'RELIABILITY_MEDIUM_THRESHOLD', 'VALUE': RELIABILITY_MEDIUM, 'DESCRIPTION': f'VARIATION_COEFFICIENT < {RELIABILITY_MEDIUM}'},
    ]

    # Статистика по причинах відхилення
    if rejected_n > 0:
        rej = rejected_drugs
        only_coverage = ((~rej['COVERAGE_CLUSTER'].isin(VALID_COVERAGE_CLUSTERS)) &
                         (rej['RELIABILITY'].isin(VALID_RELIABILITY_CLASSES))).sum()
        only_reliability = ((rej['COVERAGE_CLUSTER'].isin(VALID_COVERAGE_CLUSTERS)) &
                            (~rej['RELIABILITY'].isin(VALID_RELIABILITY_CLASSES))).sum()
        both_failed = ((~rej['COVERAGE_CLUSTER'].isin(VALID_COVERAGE_CLUSTERS)) &
                       (~rej['RELIABILITY'].isin(VALID_RELIABILITY_CLASSES))).sum()

        summary_rows.extend([
            {'METRIC': 'REJECTED_COVERAGE_ONLY', 'VALUE': only_coverage, 'DESCRIPTION': 'Відхилені лише через COVERAGE'},
            {'METRIC': 'REJECTED_RELIABILITY_ONLY', 'VALUE': only_reliability, 'DESCRIPTION': 'Відхилені лише через RELIABILITY'},
            {'METRIC': 'REJECTED_BOTH', 'VALUE': both_failed, 'DESCRIPTION': 'Відхилені через обидва критерії'},
        ])

    return pd.DataFrame(summary_rows)


# =============================================================================
# ВАЛІДАЦІЯ
# =============================================================================

def validate_results(
    drug_stats: pd.DataFrame,
    valid_drugs: pd.DataFrame,
    rejected_drugs: pd.DataFrame,
    cross_table: Dict[str, pd.DataFrame]
) -> Tuple[bool, List[str]]:
    """Валідація результатів фільтрації."""

    print("\n" + "-" * 40)
    print("Валідація результатів...")

    messages = []
    all_passed = True
    check_num = 0

    def check(name: str, condition: bool, detail: str = ""):
        nonlocal all_passed, check_num
        check_num += 1
        status = "PASSED" if condition else "FAILED"
        if not condition:
            all_passed = False
        msg = f"  [{check_num:>2}] {status}: {name}"
        if detail:
            msg += f" — {detail}"
        print(msg)
        messages.append(msg)

    total = len(drug_stats)

    # 1. valid + rejected = total
    check(
        "COMPLETENESS",
        len(valid_drugs) + len(rejected_drugs) == total,
        f"valid({len(valid_drugs)}) + rejected({len(rejected_drugs)}) = {len(valid_drugs) + len(rejected_drugs)}, expected {total}"
    )

    # 2. No duplicates in valid
    check(
        "VALID_NO_DUPLICATES",
        valid_drugs['DRUGS_ID'].nunique() == len(valid_drugs),
        f"unique={valid_drugs['DRUGS_ID'].nunique()}, total={len(valid_drugs)}"
    )

    # 3. No duplicates in rejected
    check(
        "REJECTED_NO_DUPLICATES",
        rejected_drugs['DRUGS_ID'].nunique() == len(rejected_drugs),
        f"unique={rejected_drugs['DRUGS_ID'].nunique()}, total={len(rejected_drugs)}"
    )

    # 4. No overlap
    overlap = set(valid_drugs['DRUGS_ID']) & set(rejected_drugs['DRUGS_ID'])
    check(
        "NO_OVERLAP",
        len(overlap) == 0,
        f"overlap count: {len(overlap)}"
    )

    # 5. Valid coverage criteria
    if len(valid_drugs) > 0:
        valid_cov_ok = valid_drugs['COVERAGE_CLUSTER'].isin(VALID_COVERAGE_CLUSTERS).all()
        check(
            "VALID_COVERAGE_CRITERIA",
            valid_cov_ok,
            f"all COVERAGE ∈ {sorted(VALID_COVERAGE_CLUSTERS)}"
        )
    else:
        check("VALID_COVERAGE_CRITERIA", True, "no valid drugs")

    # 6. Valid reliability criteria
    if len(valid_drugs) > 0:
        valid_rel_ok = valid_drugs['RELIABILITY'].isin(VALID_RELIABILITY_CLASSES).all()
        check(
            "VALID_RELIABILITY_CRITERIA",
            valid_rel_ok,
            f"all RELIABILITY ∈ {sorted(VALID_RELIABILITY_CLASSES)}"
        )
    else:
        check("VALID_RELIABILITY_CRITERIA", True, "no valid drugs")

    # 7. Rejected: кожен НЕ відповідає хоча б одному критерію
    if len(rejected_drugs) > 0:
        rej_cov_fail = ~rejected_drugs['COVERAGE_CLUSTER'].isin(VALID_COVERAGE_CLUSTERS)
        rej_rel_fail = ~rejected_drugs['RELIABILITY'].isin(VALID_RELIABILITY_CLASSES)
        all_have_reason = (rej_cov_fail | rej_rel_fail).all()
        check(
            "REJECTED_HAVE_REASON",
            all_have_reason,
            "кожен відхилений має хоча б одну причину"
        )
    else:
        check("REJECTED_HAVE_REASON", True, "no rejected drugs")

    # 8. Cross-table margins = total
    ct_counts = cross_table['counts']
    ct_total = ct_counts.loc['TOTAL', 'TOTAL']
    check(
        "CROSS_TABLE_TOTAL",
        ct_total == total,
        f"cross-table total={ct_total}, expected={total}"
    )

    # 9. Cross-table row sums
    for cov in COVERAGE_ORDER:
        if cov in ct_counts.index:
            row_sum = sum(ct_counts.loc[cov, col] for col in RELIABILITY_ORDER if col in ct_counts.columns)
            row_total = ct_counts.loc[cov, 'TOTAL']
            check(
                f"CROSS_TABLE_ROW_{cov}",
                row_sum == row_total,
                f"sum={row_sum}, total={row_total}"
            )

    # 10. FILTER_STATUS column
    check(
        "VALID_FILTER_STATUS",
        (valid_drugs['FILTER_STATUS'] == 'VALID').all() if len(valid_drugs) > 0 else True,
        "all FILTER_STATUS = 'VALID'"
    )

    check(
        "REJECTED_FILTER_STATUS",
        (rejected_drugs['FILTER_STATUS'] == 'REJECTED').all() if len(rejected_drugs) > 0 else True,
        "all FILTER_STATUS = 'REJECTED'"
    )

    print(f"\n  Загалом: {check_num} перевірок, {'ВСІ ПРОЙШЛИ' if all_passed else 'Є ПОМИЛКИ!'}")

    return all_passed, messages


# =============================================================================
# ЕКСПОРТ CSV
# =============================================================================

def export_to_csv(
    valid_drugs: pd.DataFrame,
    rejected_drugs: pd.DataFrame,
    filter_summary: pd.DataFrame,
    output_path: Path
) -> None:
    """Експортувати основні результати в CSV."""
    print("\n" + "-" * 40)
    print("Експорт в CSV...")

    output_path.mkdir(parents=True, exist_ok=True)

    valid_drugs.to_csv(output_path / "valid_drugs.csv", index=False)
    print(f"  valid_drugs.csv: {len(valid_drugs)} рядків")

    rejected_drugs.to_csv(output_path / "rejected_drugs.csv", index=False)
    print(f"  rejected_drugs.csv: {len(rejected_drugs)} рядків")

    filter_summary.to_csv(output_path / "filter_summary.csv", index=False)
    print(f"  filter_summary.csv: {len(filter_summary)} рядків")


# =============================================================================
# ЕКСПОРТ XLSX
# =============================================================================

def _format_drugs_xlsx(wb, df: pd.DataFrame) -> None:
    """Застосувати форматування до XLSX з препаратами."""
    from openpyxl.styles import PatternFill

    ws = wb.active

    # Кольорове маркування COVERAGE_CLUSTER
    cluster_fills = {
        'HIGH': PatternFill(start_color='C6EFCE', end_color='C6EFCE', fill_type='solid'),
        'MEDIUM': PatternFill(start_color='FFEB9C', end_color='FFEB9C', fill_type='solid'),
        'LOW': PatternFill(start_color='FFC7CE', end_color='FFC7CE', fill_type='solid'),
        'INSUFFICIENT': PatternFill(start_color='D9D9D9', end_color='D9D9D9', fill_type='solid'),
    }

    reliability_fills = {
        'HIGH': PatternFill(start_color='C6EFCE', end_color='C6EFCE', fill_type='solid'),
        'MEDIUM': PatternFill(start_color='FFEB9C', end_color='FFEB9C', fill_type='solid'),
        'LOW': PatternFill(start_color='FFC7CE', end_color='FFC7CE', fill_type='solid'),
        'SINGLE_MARKET': PatternFill(start_color='D9D9D9', end_color='D9D9D9', fill_type='solid'),
    }

    columns = list(df.columns)

    # SHARE/MEDIAN/MEAN колонки → % формат
    share_col_indices = [i + 1 for i, c in enumerate(columns)
                         if 'SHARE' in c or 'MEDIAN' in c or 'MEAN' in c or
                         'WEIGHTED_MEAN' in c or 'CI_95' in c or
                         'MIN_SHARE' in c or 'MAX_SHARE' in c or 'Q1_' in c or
                         'Q3_' in c or 'IQR_' in c or 'COVERAGE' in c.upper()
                         and c not in ('COVERAGE_CLUSTER',)]

    # Виправлення: COVERAGE_CLUSTER не повинна бути в share_col_indices
    coverage_cluster_idx = columns.index('COVERAGE_CLUSTER') + 1 if 'COVERAGE_CLUSTER' in columns else None
    share_col_indices = [i for i in share_col_indices if i != coverage_cluster_idx]

    for col_idx in share_col_indices:
        for row in range(2, ws.max_row + 1):
            cell = ws.cell(row=row, column=col_idx)
            if cell.value is not None:
                cell.number_format = '0.00%'

    # Кольори COVERAGE_CLUSTER
    if coverage_cluster_idx:
        for row in range(2, ws.max_row + 1):
            cell = ws.cell(row=row, column=coverage_cluster_idx)
            if cell.value in cluster_fills:
                cell.fill = cluster_fills[cell.value]

    # Кольори RELIABILITY
    reliability_idx = columns.index('RELIABILITY') + 1 if 'RELIABILITY' in columns else None
    if reliability_idx:
        for row in range(2, ws.max_row + 1):
            cell = ws.cell(row=row, column=reliability_idx)
            if cell.value in reliability_fills:
                cell.fill = reliability_fills[cell.value]


def export_cross_table_xlsx(
    cross_table: Dict[str, pd.DataFrame],
    total_researched: int,
    output_path: Path
) -> None:
    """Експортувати cross-table в XLSX з форматуванням."""
    from openpyxl import load_workbook
    from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
    from openpyxl.utils import get_column_letter

    xlsx_path = output_path / "cross_table_distribution.xlsx"

    with pd.ExcelWriter(xlsx_path, engine='openpyxl') as writer:
        # Sheet 1: Абсолютні значення
        cross_table['counts'].to_excel(writer, sheet_name='Counts')

        # Sheet 2: Частки (ratios)
        cross_table['ratios'].to_excel(writer, sheet_name='Ratios')

        # Sheet 3: Комбінований вигляд
        cross_table['combined'].to_excel(writer, sheet_name='Combined')

    # Форматування
    wb = load_workbook(xlsx_path)

    # --- Sheet Ratios: формат % ---
    ws_ratios = wb['Ratios']
    for row in range(2, ws_ratios.max_row + 1):
        for col in range(2, ws_ratios.max_column + 1):
            cell = ws_ratios.cell(row=row, column=col)
            if cell.value is not None:
                cell.number_format = '0.0%'

    # --- Кольорове маркування для всіх sheets ---
    cluster_fills = {
        'HIGH': PatternFill(start_color='C6EFCE', end_color='C6EFCE', fill_type='solid'),
        'MEDIUM': PatternFill(start_color='FFEB9C', end_color='FFEB9C', fill_type='solid'),
        'LOW': PatternFill(start_color='FFC7CE', end_color='FFC7CE', fill_type='solid'),
        'INSUFFICIENT': PatternFill(start_color='D9D9D9', end_color='D9D9D9', fill_type='solid'),
        'SINGLE_MARKET': PatternFill(start_color='D9D9D9', end_color='D9D9D9', fill_type='solid'),
        'TOTAL': PatternFill(start_color='BDD7EE', end_color='BDD7EE', fill_type='solid'),
    }

    bold_font = Font(bold=True)
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )

    for ws in [wb['Counts'], wb['Ratios'], wb['Combined']]:
        # Заголовки жирним
        for col in range(1, ws.max_column + 1):
            cell = ws.cell(row=1, column=col)
            cell.font = bold_font
            if cell.value in cluster_fills:
                cell.fill = cluster_fills[cell.value]

        # Рядки — кольори по назві рядка
        for row in range(2, ws.max_row + 1):
            row_label = ws.cell(row=row, column=1).value
            ws.cell(row=row, column=1).font = bold_font
            if row_label in cluster_fills:
                ws.cell(row=row, column=1).fill = cluster_fills[row_label]

            # TOTAL рядок — жирний + колір
            if row_label == 'TOTAL':
                for col in range(1, ws.max_column + 1):
                    ws.cell(row=row, column=col).font = bold_font
                    ws.cell(row=row, column=col).fill = cluster_fills['TOTAL']

        # TOTAL стовпець — жирний + колір
        total_col = None
        for col in range(1, ws.max_column + 1):
            if ws.cell(row=1, column=col).value == 'TOTAL':
                total_col = col
                break

        if total_col:
            for row in range(1, ws.max_row + 1):
                ws.cell(row=row, column=total_col).font = bold_font
                if row > 1:
                    ws.cell(row=row, column=total_col).fill = cluster_fills['TOTAL']

        # Borders
        for row in range(1, ws.max_row + 1):
            for col in range(1, ws.max_column + 1):
                ws.cell(row=row, column=col).border = thin_border

    # --- Пояснювальні нотатки під таблицями ---
    note_font = Font(italic=True, size=9, color='444444')
    note_row_offset = 2  # скільки порожніх рядків після таблиці

    for ws in [wb['Counts'], wb['Ratios'], wb['Combined']]:
        note_start_row = ws.max_row + note_row_offset
        max_col_letter = get_column_letter(ws.max_column)

        # Рядок 1: пояснення осей
        ws.merge_cells(f'A{note_start_row}:{max_col_letter}{note_start_row}')
        cell = ws.cell(row=note_start_row, column=1)
        cell.value = (
            "Рядки: COVERAGE — покриття ринків (на скількох ринках присутній препарат). "
            "Стовпці: RELIABILITY — надійність коефіцієнта субституції (на основі VARIATION_COEFFICIENT)."
        )
        cell.font = note_font
        cell.alignment = Alignment(wrap_text=True)

    # Додаткова нотатка для Ratios
    ws_ratios = wb['Ratios']
    note_row_2 = ws_ratios.max_row + 1
    max_col_letter = get_column_letter(ws_ratios.max_column)
    ws_ratios.merge_cells(f'A{note_row_2}:{max_col_letter}{note_row_2}')
    cell = ws_ratios.cell(row=note_row_2, column=1)
    cell.value = (
        f"Відсотки розраховані відносно загальної кількості досліджених препаратів ({total_researched}). "
        f"Наприклад, 17.2% означає що 87 з {total_researched} препаратів мають покриття ≥50% ринків (HIGH coverage) "
        f"та низьку варіативність показників субституції (VARIATION_COEFFICIENT < {RELIABILITY_HIGH}, HIGH reliability)."
    )
    cell.font = note_font
    cell.alignment = Alignment(wrap_text=True)

    # Додаткова нотатка для Combined
    ws_combined = wb['Combined']
    note_row_3 = ws_combined.max_row + 1
    max_col_letter = get_column_letter(ws_combined.max_column)
    ws_combined.merge_cells(f'A{note_row_3}:{max_col_letter}{note_row_3}')
    cell = ws_combined.cell(row=note_row_3, column=1)
    cell.value = (
        f"Формат: кількість (% від загальної кількості {total_researched} досліджених препаратів)."
    )
    cell.font = note_font
    cell.alignment = Alignment(wrap_text=True)

    wb.save(xlsx_path)
    print(f"  cross_table_distribution.xlsx: 3 sheets (Counts, Ratios, Combined)")


def export_pipeline_funnel_xlsx(
    funnel_df: pd.DataFrame,
    interpolation_df: pd.DataFrame,
    output_path: Path
) -> None:
    """Експортувати pipeline funnel в XLSX."""
    from openpyxl import load_workbook
    from openpyxl.styles import PatternFill, Font, Alignment, Border, Side

    xlsx_path = output_path / "pipeline_funnel.xlsx"

    with pd.ExcelWriter(xlsx_path, engine='openpyxl') as writer:
        funnel_df.to_excel(writer, sheet_name='Pipeline Funnel', index=False)
        interpolation_df.to_excel(writer, sheet_name='Interpolation', index=False)

    # Форматування
    wb = load_workbook(xlsx_path)
    bold_font = Font(bold=True)
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )

    # --- Sheet 1: Pipeline Funnel ---
    ws = wb['Pipeline Funnel']

    # Заголовки
    for col in range(1, ws.max_column + 1):
        ws.cell(row=1, column=col).font = bold_font

    # Ratio колонки → %
    ratio_cols = []
    for col in range(1, ws.max_column + 1):
        header = ws.cell(row=1, column=col).value
        if header and 'RATIO' in str(header):
            ratio_cols.append(col)

    for col_idx in ratio_cols:
        for row in range(2, ws.max_row + 1):
            cell = ws.cell(row=row, column=col_idx)
            if cell.value is not None:
                cell.number_format = '0.0%'

    # Кольорове маркування етапів
    stage_fills = {
        '1.': PatternFill(start_color='BDD7EE', end_color='BDD7EE', fill_type='solid'),
        '2.': PatternFill(start_color='C6EFCE', end_color='C6EFCE', fill_type='solid'),
        '3.': PatternFill(start_color='C6EFCE', end_color='C6EFCE', fill_type='solid'),
        '4.': PatternFill(start_color='FFC7CE', end_color='FFC7CE', fill_type='solid'),
        '5.': PatternFill(start_color='D9D9D9', end_color='D9D9D9', fill_type='solid'),
    }

    for row in range(2, ws.max_row + 1):
        stage_val = str(ws.cell(row=row, column=1).value or '')
        for prefix, fill in stage_fills.items():
            if stage_val.startswith(prefix):
                for col in range(1, ws.max_column + 1):
                    ws.cell(row=row, column=col).fill = fill
                break

    # Borders
    for row in range(1, ws.max_row + 1):
        for col in range(1, ws.max_column + 1):
            ws.cell(row=row, column=col).border = thin_border

    # --- Sheet 2: Interpolation ---
    ws2 = wb['Interpolation']
    for col in range(1, ws2.max_column + 1):
        ws2.cell(row=1, column=col).font = bold_font

    for row in range(1, ws2.max_row + 1):
        for col in range(1, ws2.max_column + 1):
            ws2.cell(row=row, column=col).border = thin_border

    wb.save(xlsx_path)
    print(f"  pipeline_funnel.xlsx: 2 sheets (Pipeline Funnel, Interpolation)")


def export_to_xlsx(
    valid_drugs: pd.DataFrame,
    rejected_drugs: pd.DataFrame,
    cross_table: Dict[str, pd.DataFrame],
    funnel_df: pd.DataFrame,
    interpolation_df: pd.DataFrame,
    output_path: Path
) -> None:
    """Експортувати всі бізнес-звіти в XLSX."""
    print("\n" + "-" * 40)
    print("Експорт в XLSX для бізнесу...")

    output_path.mkdir(parents=True, exist_ok=True)

    try:
        from openpyxl import load_workbook

        # --- valid_drugs.xlsx ---
        valid_path = output_path / "valid_drugs.xlsx"
        valid_drugs.to_excel(valid_path, index=False, sheet_name="Valid Drugs")
        wb = load_workbook(valid_path)
        _format_drugs_xlsx(wb, valid_drugs)
        wb.save(valid_path)
        print(f"  valid_drugs.xlsx: {len(valid_drugs)} рядків")

        # --- rejected_drugs.xlsx ---
        rejected_path = output_path / "rejected_drugs.xlsx"
        rejected_drugs.to_excel(rejected_path, index=False, sheet_name="Rejected Drugs")
        wb = load_workbook(rejected_path)
        _format_drugs_xlsx(wb, rejected_drugs)
        wb.save(rejected_path)
        print(f"  rejected_drugs.xlsx: {len(rejected_drugs)} рядків")

        # --- cross_table_distribution.xlsx ---
        total_researched = len(valid_drugs) + len(rejected_drugs)
        export_cross_table_xlsx(cross_table, total_researched, output_path)

        # --- pipeline_funnel.xlsx ---
        export_pipeline_funnel_xlsx(funnel_df, interpolation_df, output_path)

    except ImportError:
        print("  ПОМИЛКА: openpyxl не встановлено (pip install openpyxl)")
    except Exception as e:
        print(f"  ПОМИЛКА при експорті XLSX: {e}")
        import traceback
        traceback.print_exc()


# =============================================================================
# VALIDATION REPORT
# =============================================================================

def create_validation_report(
    all_passed: bool,
    messages: List[str],
    output_path: Path
) -> None:
    """Створити текстовий звіт валідації."""
    report_path = output_path / "validation_report.txt"

    with open(report_path, 'w', encoding='utf-8') as f:
        f.write("=" * 70 + "\n")
        f.write("VALIDATION REPORT — 02_02_valid_data_filter\n")
        f.write(f"Generated: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}\n")
        f.write("=" * 70 + "\n\n")

        f.write(f"Filter: Scenario A (strict)\n")
        f.write(f"  COVERAGE_CLUSTER ∈ {{HIGH, MEDIUM}}\n")
        f.write(f"  RELIABILITY ∈ {{HIGH, MEDIUM}}\n\n")

        f.write(f"Thresholds from config:\n")
        f.write(f"  COVERAGE_HIGH = {COVERAGE_HIGH} (≥{COVERAGE_HIGH:.0%})\n")
        f.write(f"  COVERAGE_MEDIUM = {COVERAGE_MEDIUM} (≥{COVERAGE_MEDIUM:.0%})\n")
        f.write(f"  RELIABILITY_HIGH = {RELIABILITY_HIGH} (VARIATION_COEFFICIENT < {RELIABILITY_HIGH})\n")
        f.write(f"  RELIABILITY_MEDIUM = {RELIABILITY_MEDIUM} (VARIATION_COEFFICIENT < {RELIABILITY_MEDIUM})\n\n")

        f.write("Validation checks:\n")
        for msg in messages:
            f.write(msg + "\n")

        f.write(f"\nOverall: {'ALL PASSED' if all_passed else 'SOME FAILED'}\n")

    print(f"\nValidation report saved: {report_path}")


# =============================================================================
# MAIN
# =============================================================================

def main():
    """Головна функція."""
    print("=" * 70)
    print("PHASE 2, STEP 2.2: VALID DATA FILTER (SCENARIO A — STRICT)")
    print(f"Started: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}")
    print("=" * 70)

    # 1. Завантаження даних
    drug_stats = load_drug_statistics()
    all_drugs = load_all_drugs_list()
    raw_drugs_count = count_raw_unique_drugs()

    total_markets = drug_stats['MARKET_COUNT_TOTAL'].max() if 'MARKET_COUNT_TOTAL' in drug_stats.columns else 97

    # 2. Фільтрація
    valid_drugs, rejected_drugs = apply_filter(drug_stats)

    # 3. Cross-table
    cross_table = create_cross_table(drug_stats)

    # 4. Pipeline funnel
    funnel_df, interpolation_df = create_pipeline_funnel(
        raw_drugs_count=raw_drugs_count,
        all_drugs_count=len(all_drugs),
        researched_count=len(drug_stats),
        valid_count=len(valid_drugs),
        total_markets=total_markets
    )

    # 5. Filter summary
    filter_summary = create_filter_summary(drug_stats, valid_drugs, rejected_drugs)

    # 6. Валідація
    all_passed, messages = validate_results(drug_stats, valid_drugs, rejected_drugs, cross_table)

    # 7. Експорт CSV
    export_to_csv(valid_drugs, rejected_drugs, filter_summary, OUTPUT_BASE_PATH)

    # 8. Експорт XLSX
    export_to_xlsx(
        valid_drugs, rejected_drugs,
        cross_table, funnel_df, interpolation_df,
        OUTPUT_BUSINESS_PATH
    )

    # 9. Validation report
    create_validation_report(all_passed, messages, OUTPUT_BASE_PATH)

    # Summary
    print("\n" + "=" * 70)
    print("SUMMARY")
    print("=" * 70)
    print(f"  Input: {len(drug_stats)} drugs from drug_statistics.csv")
    print(f"  Valid drugs: {len(valid_drugs)} ({len(valid_drugs)/len(drug_stats):.1%})")
    print(f"  Rejected drugs: {len(rejected_drugs)} ({len(rejected_drugs)/len(drug_stats):.1%})")
    print(f"  Validation: {'PASSED' if all_passed else 'FAILED'}")
    print(f"\nOutput folder: {OUTPUT_BASE_PATH}")
    print(f"Business reports: {OUTPUT_BUSINESS_PATH}")
    print(f"\nFinished: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}")
    print("=" * 70)


if __name__ == "__main__":
    main()
