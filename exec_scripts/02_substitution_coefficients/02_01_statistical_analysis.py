"""
02_01_statistical_analysis.py - Статистичний аналіз коефіцієнтів субституції

Phase 2, Step 2.2 part 1: Statistical Analysis

Вхідні дані:
- results/cross_market_data/market_substitution_{CLIENT_ID}/sub_coef_{CLIENT_ID}.csv (97 файлів)
- results/substitution_research/01_preparation/researched_drugs_list.csv (coverage кластери)

Вихідні дані:
- results/substitution_research/02_statistics_and_filter/02_01_statistical_analysis/drug_statistics.csv
- results/substitution_research/02_statistics_and_filter/02_01_statistical_analysis/drug_distribution.csv
- results/substitution_research/02_statistics_and_filter/02_01_statistical_analysis/flat_bi_export.csv
- results/substitution_research/02_statistics_and_filter/02_01_statistical_analysis/validation_report.txt
- results/substitution_research/02_statistics_and_filter/02_01_statistical_analysis/stat_business_reports/*.xlsx

Використання:
    python exec_scripts/02_substitution_coefficients/02_01_statistical_analysis.py
"""

import pandas as pd
import numpy as np
import sys
from pathlib import Path
from datetime import datetime
from typing import List, Dict, Tuple

# Додаємо шлях до project_core
SCRIPT_PATH = Path(__file__).resolve()
PROJECT_ROOT = SCRIPT_PATH.parent.parent.parent
sys.path.insert(0, str(PROJECT_ROOT))

from project_core.data_config.paths_config import RESULTS_PATH
from project_core.sub_coef_config.coverage_thresholds import get_coverage_cluster
from project_core.sub_coef_config.reliability_thresholds import get_reliability_class


# =============================================================================
# КОНСТАНТИ
# =============================================================================

# Вхідні шляхи
CROSS_MARKET_PATH = RESULTS_PATH / "cross_market_data"
PREPARATION_PATH = RESULTS_PATH / "substitution_research" / "01_preparation"

# Вихідні шляхи
OUTPUT_BASE_PATH = RESULTS_PATH / "substitution_research" / "02_statistics_and_filter" / "02_01_statistical_analysis"
OUTPUT_BUSINESS_PATH = OUTPUT_BASE_PATH / "stat_business_reports"

# Діапазони для гістограми (крок 10%)
BIN_EDGES = [0.0, 0.1, 0.2, 0.3, 0.4, 0.5, 0.6, 0.7, 0.8, 0.9, 1.0]
BIN_LABELS = [
    'BIN_0_10', 'BIN_10_20', 'BIN_20_30', 'BIN_30_40', 'BIN_40_50',
    'BIN_50_60', 'BIN_60_70', 'BIN_70_80', 'BIN_80_90', 'BIN_90_100'
]

# IQR множник для визначення outliers
IQR_MULTIPLIER = 1.5


# =============================================================================
# ЗАВАНТАЖЕННЯ ДАНИХ
# =============================================================================

def load_all_sub_coef() -> Tuple[pd.DataFrame, int]:
    """
    Завантажити та об'єднати всі sub_coef_*.csv файли.

    Returns:
        Tuple: (concat DataFrame, кількість завантажених файлів)
    """
    print("\n" + "=" * 60)
    print("ЗАВАНТАЖЕННЯ SUB_COEF ФАЙЛІВ")
    print("=" * 60)

    sub_coef_files = sorted(CROSS_MARKET_PATH.glob("market_substitution_*/sub_coef_*.csv"))

    if not sub_coef_files:
        raise ValueError(f"Не знайдено sub_coef файлів у {CROSS_MARKET_PATH}")

    print(f"Знайдено файлів: {len(sub_coef_files)}")

    all_data = []
    for f in sub_coef_files:
        try:
            df = pd.read_csv(f)
            all_data.append(df)
        except Exception as e:
            print(f"  ПОМИЛКА при читанні {f.name}: {e}")

    combined = pd.concat(all_data, ignore_index=True)

    print(f"Всього записів: {len(combined)}")
    print(f"Унікальних препаратів: {combined['DRUGS_ID'].nunique()}")
    print(f"Унікальних ринків: {combined['CLIENT_ID'].nunique()}")

    return combined, len(sub_coef_files)


def load_researched_drugs_list() -> pd.DataFrame:
    """
    Завантажити список досліджених препаратів з coverage кластерами.

    Returns:
        DataFrame з coverage metadata
    """
    path = PREPARATION_PATH / "researched_drugs_list.csv"
    if not path.exists():
        raise FileNotFoundError(f"Не знайдено: {path}")

    df = pd.read_csv(path)
    print(f"Завантажено researched_drugs_list: {len(df)} препаратів")
    return df


# =============================================================================
# IQR ФІЛЬТРАЦІЯ OUTLIERS
# =============================================================================

def mark_outliers(df: pd.DataFrame) -> pd.DataFrame:
    """
    Позначити outliers за IQR методом per DRUGS_ID.

    Алгоритм:
    - Per DRUGS_ID: розрахунок Q1, Q3, IQR для SHARE_INTERNAL
    - Outlier = значення за межами [Q1 - 1.5*IQR, Q3 + 1.5*IQR]
    - Препарати з одним спостереженням → SINGLE_OBSERVATION

    Args:
        df: DataFrame з усіма sub_coef даними

    Returns:
        DataFrame з доданими колонками IS_OUTLIER, SINGLE_OBSERVATION
    """
    print("\n" + "=" * 60)
    print("IQR ФІЛЬТРАЦІЯ OUTLIERS")
    print("=" * 60)

    result = df.copy()
    result['IS_OUTLIER'] = False
    result['SINGLE_OBSERVATION'] = False

    # Per-drug IQR розрахунок (vectorized через groupby + transform)
    grouped = result.groupby('DRUGS_ID')['SHARE_INTERNAL']
    q1 = grouped.transform('quantile', 0.25)
    q3 = grouped.transform('quantile', 0.75)
    iqr = q3 - q1
    lower_bound = q1 - IQR_MULTIPLIER * iqr
    upper_bound = q3 + IQR_MULTIPLIER * iqr

    # Позначаємо outliers
    result['IS_OUTLIER'] = (result['SHARE_INTERNAL'] < lower_bound) | \
                           (result['SHARE_INTERNAL'] > upper_bound)

    # Препарати з одним спостереженням
    drug_counts = grouped.transform('count')
    result['SINGLE_OBSERVATION'] = drug_counts == 1

    # Статистика
    total_outliers = result['IS_OUTLIER'].sum()
    single_obs = result['SINGLE_OBSERVATION'].sum()
    total = len(result)

    print(f"  Всього записів: {total}")
    print(f"  Outliers (IQR): {total_outliers} ({total_outliers/total*100:.1f}%)")
    print(f"  Single observation: {single_obs}")
    print(f"  Clean records: {total - total_outliers} ({(total - total_outliers)/total*100:.1f}%)")

    return result


# =============================================================================
# DRUG STATISTICS
# =============================================================================

def create_drug_statistics(
    df: pd.DataFrame,
    researched_drugs: pd.DataFrame,
    total_markets: int
) -> pd.DataFrame:
    """
    Розрахувати статистику per DRUGS_ID (без outliers).

    Args:
        df: DataFrame з IS_OUTLIER флагом
        researched_drugs: DataFrame з coverage metadata
        total_markets: Загальна кількість ринків

    Returns:
        DataFrame для drug_statistics.csv
    """
    print("\n" + "=" * 60)
    print("РОЗРАХУНОК DRUG STATISTICS")
    print("=" * 60)

    # Фільтруємо outliers для статистики
    clean = df[~df['IS_OUTLIER']].copy()

    # Групуємо по DRUGS_ID
    stats = clean.groupby('DRUGS_ID').agg(
        MEDIAN_SHARE_INTERNAL=('SHARE_INTERNAL', 'median'),
        MEAN_SHARE_INTERNAL=('SHARE_INTERNAL', 'mean'),
        STD_SHARE_INTERNAL=('SHARE_INTERNAL', 'std'),
        MIN_SHARE_INTERNAL=('SHARE_INTERNAL', 'min'),
        MAX_SHARE_INTERNAL=('SHARE_INTERNAL', 'max'),
        Q1_SHARE_INTERNAL=('SHARE_INTERNAL', lambda x: x.quantile(0.25)),
        Q3_SHARE_INTERNAL=('SHARE_INTERNAL', lambda x: x.quantile(0.75)),
        MARKET_COUNT_CLEAN=('CLIENT_ID', 'nunique'),
        TOTAL_EVENTS=('EVENTS_COUNT', 'sum'),
        TOTAL_INTERNAL_LIFT=('INTERNAL_LIFT', 'sum'),
    ).reset_index()

    # IQR та VARIATION_COEFFICIENT
    stats['IQR_SHARE_INTERNAL'] = stats['Q3_SHARE_INTERNAL'] - stats['Q1_SHARE_INTERNAL']
    stats['VARIATION_COEFFICIENT'] = np.where(
        stats['MEAN_SHARE_INTERNAL'] > 0,
        stats['STD_SHARE_INTERNAL'] / stats['MEAN_SHARE_INTERNAL'],
        np.nan
    )

    # WEIGHTED_MEAN — зважене середнє по INTERNAL_LIFT (методологія секція 3.1)
    def _weighted_mean(group):
        total_lift = group['INTERNAL_LIFT'].sum()
        if total_lift > 0:
            return (group['SHARE_INTERNAL'] * group['INTERNAL_LIFT']).sum() / total_lift
        else:
            return group['SHARE_INTERNAL'].mean()  # fallback: simple mean (edge case 7.3)

    weighted = clean.groupby('DRUGS_ID').apply(_weighted_mean, include_groups=False).reset_index()
    weighted.columns = ['DRUGS_ID', 'WEIGHTED_MEAN_SHARE']
    stats = stats.merge(weighted, on='DRUGS_ID', how='left')

    # 95% CI (методологія секція 5)
    stats['CI_95_LOWER'] = np.where(
        stats['MARKET_COUNT_CLEAN'] >= 2,
        (stats['MEAN_SHARE_INTERNAL'] - 1.96 * stats['STD_SHARE_INTERNAL'] / np.sqrt(stats['MARKET_COUNT_CLEAN'])).clip(0, 1),
        np.nan
    )
    stats['CI_95_UPPER'] = np.where(
        stats['MARKET_COUNT_CLEAN'] >= 2,
        (stats['MEAN_SHARE_INTERNAL'] + 1.96 * stats['STD_SHARE_INTERNAL'] / np.sqrt(stats['MARKET_COUNT_CLEAN'])).clip(0, 1),
        np.nan
    )

    # RELIABILITY classification (пороги з project_core/sub_coef_config/reliability_thresholds.py)
    stats['RELIABILITY'] = stats.apply(
        lambda row: get_reliability_class(
            variation_coefficient=row['VARIATION_COEFFICIENT'] if pd.notna(row['VARIATION_COEFFICIENT']) else float('nan'),
            market_count_clean=row['MARKET_COUNT_CLEAN']
        ),
        axis=1
    )

    # Кількість outliers per drug
    outlier_counts = df[df['IS_OUTLIER']].groupby('DRUGS_ID').size().reset_index(name='OUTLIERS_COUNT')
    stats = stats.merge(outlier_counts, on='DRUGS_ID', how='left')
    stats['OUTLIERS_COUNT'] = stats['OUTLIERS_COUNT'].fillna(0).astype(int)

    # Загальний MARKET_COUNT (включно з outliers)
    total_counts = df.groupby('DRUGS_ID')['CLIENT_ID'].nunique().reset_index(name='MARKET_COUNT_TOTAL')
    stats = stats.merge(total_counts, on='DRUGS_ID', how='left')

    # Join з метаданими
    meta_cols = ['DRUGS_ID', 'DRUGS_NAME', 'INN_ID', 'INN_NAME', 'NFC1_ID',
                 'MARKET_COVERAGE', 'COVERAGE_CLUSTER']
    meta = researched_drugs[meta_cols].copy()
    stats = stats.merge(meta, on='DRUGS_ID', how='left')

    # Сортування по median DESC
    stats = stats.sort_values('MEDIAN_SHARE_INTERNAL', ascending=False).reset_index(drop=True)

    # Порядок колонок
    col_order = [
        'DRUGS_ID', 'DRUGS_NAME', 'INN_ID', 'INN_NAME', 'NFC1_ID',
        'COVERAGE_CLUSTER', 'RELIABILITY',
        'MARKET_COUNT_TOTAL', 'MARKET_COUNT_CLEAN', 'OUTLIERS_COUNT',
        'MEDIAN_SHARE_INTERNAL', 'MEAN_SHARE_INTERNAL', 'WEIGHTED_MEAN_SHARE',
        'STD_SHARE_INTERNAL', 'VARIATION_COEFFICIENT',
        'CI_95_LOWER', 'CI_95_UPPER',
        'MIN_SHARE_INTERNAL', 'Q1_SHARE_INTERNAL',
        'Q3_SHARE_INTERNAL', 'MAX_SHARE_INTERNAL', 'IQR_SHARE_INTERNAL',
        'MARKET_COVERAGE', 'TOTAL_EVENTS', 'TOTAL_INTERNAL_LIFT'
    ]
    stats = stats[col_order]

    print(f"  Препаратів: {len(stats)}")
    print(f"  Median SHARE_INTERNAL (overall): {stats['MEDIAN_SHARE_INTERNAL'].median():.4f}")
    print(f"  VARIATION_COEFFICIENT (overall median): {stats['VARIATION_COEFFICIENT'].median():.4f}")

    return stats


# =============================================================================
# DRUG DISTRIBUTION
# =============================================================================

def create_drug_distribution(df: pd.DataFrame, researched_drugs: pd.DataFrame) -> pd.DataFrame:
    """
    Розподіл SHARE_INTERNAL по діапазонах 10% per DRUGS_ID (без outliers).

    Args:
        df: DataFrame з IS_OUTLIER флагом
        researched_drugs: DataFrame з coverage metadata

    Returns:
        DataFrame для drug_distribution.csv
    """
    print("\n" + "=" * 60)
    print("РОЗРАХУНОК DRUG DISTRIBUTION (діапазони 10%)")
    print("=" * 60)

    clean = df[~df['IS_OUTLIER']].copy()

    # Bins per drug — vectorized
    clean['BIN'] = pd.cut(
        clean['SHARE_INTERNAL'],
        bins=BIN_EDGES,
        labels=BIN_LABELS,
        include_lowest=True,
        right=True
    )

    # Pivot: DRUGS_ID x BIN → count
    pivot = clean.groupby(['DRUGS_ID', 'BIN'], observed=False).size().unstack(fill_value=0)
    pivot.columns = pivot.columns.astype(str)
    pivot = pivot.reset_index()

    # Додаємо TOTAL
    pivot['TOTAL'] = pivot[BIN_LABELS].sum(axis=1)

    # Join з метаданими
    meta_cols = ['DRUGS_ID', 'DRUGS_NAME', 'INN_ID', 'INN_NAME']
    meta = researched_drugs[meta_cols].drop_duplicates(subset=['DRUGS_ID'])
    result = meta.merge(pivot, on='DRUGS_ID', how='right')

    # Сортування по TOTAL DESC
    result = result.sort_values('TOTAL', ascending=False).reset_index(drop=True)

    print(f"  Препаратів: {len(result)}")
    print(f"  Діапазони: {len(BIN_LABELS)}")

    # Топ-3 найзаповненіших діапазони (загалом)
    bin_totals = result[BIN_LABELS].sum()
    top_bins = bin_totals.sort_values(ascending=False).head(3)
    print(f"  Топ-3 діапазони по кількості спостережень:")
    for bin_name, count in top_bins.items():
        pct_label = bin_name.replace('BIN_', '').replace('_', '-') + '%'
        print(f"    {pct_label}: {int(count)} спостережень")

    return result


# =============================================================================
# FLAT BI EXPORT
# =============================================================================

def create_flat_bi_export(df: pd.DataFrame) -> pd.DataFrame:
    """
    Створити flat export для Power BI (long format, без пустих полів).

    Args:
        df: DataFrame з IS_OUTLIER та SINGLE_OBSERVATION флагами

    Returns:
        DataFrame для flat_bi_export.csv
    """
    print("\n" + "=" * 60)
    print("ФОРМУВАННЯ FLAT BI EXPORT")
    print("=" * 60)

    # Вибираємо колонки для BI
    bi_cols = [
        'CLIENT_ID', 'DRUGS_ID', 'DRUGS_NAME', 'INN_ID', 'INN_NAME', 'NFC1_ID',
        'EVENTS_COUNT', 'TOTAL_STOCKOUT_WEEKS',
        'INTERNAL_LIFT', 'LOST_SALES', 'TOTAL_EFFECT',
        'SHARE_INTERNAL', 'SHARE_LOST',
        'SHARE_SAME_NFC1', 'SHARE_DIFF_NFC1',
        'CLASSIFICATION',
        'IS_OUTLIER', 'SINGLE_OBSERVATION'
    ]

    # Перевіряємо які колонки доступні
    available_cols = [c for c in bi_cols if c in df.columns]
    result = df[available_cols].copy()

    # Сортування для зручності BI
    result = result.sort_values(['DRUGS_ID', 'CLIENT_ID']).reset_index(drop=True)

    print(f"  Записів: {len(result)}")
    print(f"  Колонок: {len(result.columns)}")
    print(f"  Унікальних препаратів: {result['DRUGS_ID'].nunique()}")
    print(f"  Унікальних ринків: {result['CLIENT_ID'].nunique()}")

    return result


# =============================================================================
# ВАЛІДАЦІЯ
# =============================================================================

def validate_results(
    raw_data: pd.DataFrame,
    drug_stats: pd.DataFrame,
    drug_dist: pd.DataFrame,
    flat_export: pd.DataFrame,
    researched_drugs: pd.DataFrame,
    total_files: int
) -> Tuple[bool, List[str]]:
    """
    Валідація всіх результатів.

    Returns:
        Tuple: (всі тести пройшли, список повідомлень)
    """
    print("\n" + "=" * 60)
    print("ВАЛІДАЦІЯ РЕЗУЛЬТАТІВ")
    print("=" * 60)

    messages = []
    all_passed = True

    # 1. Кількість препаратів
    stats_count = len(drug_stats)
    researched_count = len(researched_drugs)
    if stats_count == researched_count:
        messages.append(f"[OK] DRUGS_COUNT: {stats_count} = {researched_count} (researched_drugs_list)")
    else:
        messages.append(f"[WARN] DRUGS_COUNT: stats={stats_count}, researched={researched_count}")

    # 2. MEDIAN в [0, 1]
    median_values = drug_stats['MEDIAN_SHARE_INTERNAL']
    out_of_range = ((median_values < 0) | (median_values > 1)).sum()
    if out_of_range == 0:
        messages.append(f"[OK] MEDIAN_RANGE: all {len(median_values)} values in [0, 1]")
    else:
        messages.append(f"[FAIL] MEDIAN_RANGE: {out_of_range} values out of [0, 1]")
        all_passed = False

    # 3. MARKET_COUNT <= total_files
    over_count = (drug_stats['MARKET_COUNT_TOTAL'] > total_files).sum()
    if over_count == 0:
        messages.append(f"[OK] MARKET_COUNT: all <= {total_files}")
    else:
        messages.append(f"[FAIL] MARKET_COUNT: {over_count} drugs exceed {total_files} markets")
        all_passed = False

    # 4. Distribution bins sum = TOTAL per drug
    if len(drug_dist) > 0:
        bin_sum = drug_dist[BIN_LABELS].sum(axis=1)
        mismatch = (bin_sum != drug_dist['TOTAL']).sum()
        if mismatch == 0:
            messages.append(f"[OK] DISTRIBUTION_SUM: all {len(drug_dist)} drugs bins sum = TOTAL")
        else:
            messages.append(f"[FAIL] DISTRIBUTION_SUM: {mismatch} drugs have bin sum != TOTAL")
            all_passed = False

    # 5. Flat export records = raw data records
    if len(flat_export) == len(raw_data):
        messages.append(f"[OK] FLAT_EXPORT: {len(flat_export)} records = raw data")
    else:
        messages.append(f"[FAIL] FLAT_EXPORT: {len(flat_export)} != {len(raw_data)} raw records")
        all_passed = False

    # 6. Cross-check MARKET_COUNT_TOTAL з researched_drugs_list
    for _, row in drug_stats.head(5).iterrows():
        drug_id = row['DRUGS_ID']
        ref = researched_drugs[researched_drugs['DRUGS_ID'] == drug_id]
        if len(ref) > 0:
            ref_count = ref.iloc[0]['MARKET_COUNT']
            actual_count = row['MARKET_COUNT_TOTAL']
            if actual_count == ref_count:
                messages.append(f"[OK] CROSS_CHECK drug {drug_id}: {actual_count} markets")
            else:
                messages.append(f"[FAIL] CROSS_CHECK drug {drug_id}: stats={actual_count}, ref={ref_count}")
                all_passed = False

    # 7. VARIATION_COEFFICIENT >= 0 (де визначено)
    valid_vc = drug_stats['VARIATION_COEFFICIENT'].dropna()
    negative_vc = (valid_vc < 0).sum()
    if negative_vc == 0:
        messages.append(f"[OK] VARIATION_COEFFICIENT_RANGE: all {len(valid_vc)} values >= 0")
    else:
        messages.append(f"[FAIL] VARIATION_COEFFICIENT_RANGE: {negative_vc} negative values")
        all_passed = False

    # 8. WEIGHTED_MEAN in [0, 1]
    wm = drug_stats['WEIGHTED_MEAN_SHARE'].dropna()
    wm_out = ((wm < 0) | (wm > 1)).sum()
    if wm_out == 0:
        messages.append(f"[OK] WEIGHTED_MEAN_RANGE: all {len(wm)} values in [0, 1]")
    else:
        messages.append(f"[FAIL] WEIGHTED_MEAN_RANGE: {wm_out} values out of [0, 1]")
        all_passed = False

    # 9. CI logic: CI_LOWER <= MEAN <= CI_UPPER
    ci_valid = drug_stats.dropna(subset=['CI_95_LOWER', 'CI_95_UPPER'])
    ci_violation = (
        (ci_valid['CI_95_LOWER'] > ci_valid['MEAN_SHARE_INTERNAL'] + 1e-9) |
        (ci_valid['CI_95_UPPER'] < ci_valid['MEAN_SHARE_INTERNAL'] - 1e-9)
    ).sum()
    if ci_violation == 0:
        messages.append(f"[OK] CI_LOGIC: all {len(ci_valid)} CI contain MEAN")
    else:
        messages.append(f"[FAIL] CI_LOGIC: {ci_violation} drugs CI does not contain MEAN")
        all_passed = False

    # 10. MIN <= MEAN <= MAX
    valid_range = drug_stats.dropna(subset=['MEAN_SHARE_INTERNAL'])
    range_violation = (
        (valid_range['MIN_SHARE_INTERNAL'] > valid_range['MEAN_SHARE_INTERNAL'] + 1e-9) |
        (valid_range['MAX_SHARE_INTERNAL'] < valid_range['MEAN_SHARE_INTERNAL'] - 1e-9)
    ).sum()
    if range_violation == 0:
        messages.append(f"[OK] MIN_MEAN_MAX: all {len(valid_range)} drugs MIN <= MEAN <= MAX")
    else:
        messages.append(f"[FAIL] MIN_MEAN_MAX: {range_violation} drugs violate MIN <= MEAN <= MAX")
        all_passed = False

    for msg in messages:
        print(f"  {msg}")

    status = "PASSED" if all_passed else "FAILED"
    print(f"\nЗАГАЛЬНИЙ СТАТУС: {status}")

    return all_passed, messages


def create_validation_report(all_passed: bool, messages: List[str], output_path: Path) -> None:
    """Створити текстовий файл валідації."""
    report_path = output_path / "validation_report.txt"

    with open(report_path, 'w', encoding='utf-8') as f:
        f.write("=" * 70 + "\n")
        f.write("VALIDATION REPORT - Phase 2 Step 2: Statistical Analysis\n")
        f.write(f"Generated: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}\n")
        f.write("=" * 70 + "\n\n")

        status = "PASSED" if all_passed else "FAILED"
        f.write(f"OVERALL STATUS: {status}\n\n")

        f.write("-" * 40 + "\n")
        f.write("VALIDATION CHECKS:\n")
        f.write("-" * 40 + "\n")

        for msg in messages:
            f.write(f"{msg}\n")

        f.write("\n" + "=" * 70 + "\n")

    print(f"\nValidation report saved: {report_path}")


# =============================================================================
# ЕКСПОРТ
# =============================================================================

def export_to_csv(
    drug_stats: pd.DataFrame,
    drug_dist: pd.DataFrame,
    flat_export: pd.DataFrame,
    output_path: Path
) -> None:
    """Експортувати всі дані в CSV."""
    print("\n" + "-" * 40)
    print("Експорт в CSV...")

    output_path.mkdir(parents=True, exist_ok=True)

    drug_stats.to_csv(output_path / "drug_statistics.csv", index=False)
    print(f"  drug_statistics.csv: {len(drug_stats)} рядків, {len(drug_stats.columns)} колонок")

    drug_dist.to_csv(output_path / "drug_distribution.csv", index=False)
    print(f"  drug_distribution.csv: {len(drug_dist)} рядків")

    flat_export.to_csv(output_path / "flat_bi_export.csv", index=False)
    print(f"  flat_bi_export.csv: {len(flat_export)} рядків, {len(flat_export.columns)} колонок")


def export_to_xlsx(
    drug_stats: pd.DataFrame,
    drug_dist: pd.DataFrame,
    output_path: Path
) -> None:
    """Експортувати дані в Excel з форматуванням."""
    print("\n" + "-" * 40)
    print("Експорт в XLSX для бізнесу...")

    output_path.mkdir(parents=True, exist_ok=True)

    try:
        from openpyxl import load_workbook
        from openpyxl.styles import PatternFill, Font, Alignment, numbers

        # --- drug_statistics.xlsx ---
        stats_path = output_path / "drug_statistics.xlsx"
        drug_stats.to_excel(stats_path, index=False, sheet_name="Drug Statistics")

        wb = load_workbook(stats_path)
        ws = wb.active

        # Визначаємо колонки з SHARE значеннями для % формату
        share_cols = [i + 1 for i, c in enumerate(drug_stats.columns)
                      if 'SHARE' in c or 'MEDIAN' in c or 'MEAN' in c or
                      'WEIGHTED_MEAN' in c or 'CI_95' in c or
                      'MIN_SHARE' in c or 'MAX_SHARE' in c or 'Q1_' in c or
                      'Q3_' in c or 'IQR_' in c or 'COVERAGE' in c]

        for col_idx in share_cols:
            for row in range(2, ws.max_row + 1):
                cell = ws.cell(row=row, column=col_idx)
                if cell.value is not None:
                    cell.number_format = '0.00%'

        # Кольорове маркування COVERAGE_CLUSTER
        cluster_col_idx = None
        for i, c in enumerate(drug_stats.columns):
            if c == 'COVERAGE_CLUSTER':
                cluster_col_idx = i + 1
                break

        if cluster_col_idx:
            fills = {
                'HIGH': PatternFill(start_color='C6EFCE', end_color='C6EFCE', fill_type='solid'),
                'MEDIUM': PatternFill(start_color='FFEB9C', end_color='FFEB9C', fill_type='solid'),
                'LOW': PatternFill(start_color='FFC7CE', end_color='FFC7CE', fill_type='solid'),
                'INSUFFICIENT': PatternFill(start_color='D9D9D9', end_color='D9D9D9', fill_type='solid'),
            }
            for row in range(2, ws.max_row + 1):
                cell = ws.cell(row=row, column=cluster_col_idx)
                if cell.value in fills:
                    cell.fill = fills[cell.value]

        # Кольорове маркування RELIABILITY
        reliability_col_idx = None
        for i, c in enumerate(drug_stats.columns):
            if c == 'RELIABILITY':
                reliability_col_idx = i + 1
                break

        if reliability_col_idx:
            rel_fills = {
                'HIGH': PatternFill(start_color='C6EFCE', end_color='C6EFCE', fill_type='solid'),
                'MEDIUM': PatternFill(start_color='FFEB9C', end_color='FFEB9C', fill_type='solid'),
                'LOW': PatternFill(start_color='FFC7CE', end_color='FFC7CE', fill_type='solid'),
                'SINGLE_MARKET': PatternFill(start_color='D9D9D9', end_color='D9D9D9', fill_type='solid'),
            }
            for row in range(2, ws.max_row + 1):
                cell = ws.cell(row=row, column=reliability_col_idx)
                if cell.value in rel_fills:
                    cell.fill = rel_fills[cell.value]

        wb.save(stats_path)
        print(f"  drug_statistics.xlsx: {len(drug_stats)} рядків (SHARE як %, cluster/reliability кольори)")

        # --- drug_distribution.xlsx ---
        dist_path = output_path / "drug_distribution.xlsx"
        drug_dist.to_excel(dist_path, index=False, sheet_name="Distribution")

        wb = load_workbook(dist_path)
        ws = wb.active

        # Conditional formatting (heatmap) на bin колонках
        bin_col_indices = [i + 1 for i, c in enumerate(drug_dist.columns) if c.startswith('BIN_')]

        # Знаходимо max значення для масштабування кольору
        max_bin_value = drug_dist[BIN_LABELS].max().max() if len(drug_dist) > 0 else 1

        for col_idx in bin_col_indices:
            for row in range(2, ws.max_row + 1):
                cell = ws.cell(row=row, column=col_idx)
                if cell.value is not None and cell.value > 0:
                    # Інтенсивність кольору пропорційна значенню
                    intensity = min(cell.value / max(max_bin_value, 1), 1.0)
                    # Від білого (0) до зеленого (max)
                    green = int(200 + (1 - intensity) * 55)
                    red = int(255 - intensity * 100)
                    blue = int(255 - intensity * 100)
                    hex_color = f'{red:02X}{green:02X}{blue:02X}'
                    cell.fill = PatternFill(start_color=hex_color, end_color=hex_color, fill_type='solid')

        wb.save(dist_path)
        print(f"  drug_distribution.xlsx: {len(drug_dist)} рядків (heatmap bins)")

    except ImportError:
        print("  ПОМИЛКА: openpyxl не встановлено (pip install openpyxl)")
    except Exception as e:
        print(f"  ПОМИЛКА при експорті XLSX: {e}")


# =============================================================================
# MAIN
# =============================================================================

def main():
    """Головна функція."""
    print("=" * 70)
    print("PHASE 2, STEP 2: STATISTICAL ANALYSIS")
    print(f"Started: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}")
    print("=" * 70)

    # 1. Завантаження даних
    raw_data, total_files = load_all_sub_coef()
    researched_drugs = load_researched_drugs_list()

    # 2. IQR фільтрація
    data_with_flags = mark_outliers(raw_data)

    # 3. Drug statistics
    drug_stats = create_drug_statistics(data_with_flags, researched_drugs, total_files)

    # 4. Drug distribution
    drug_dist = create_drug_distribution(data_with_flags, researched_drugs)

    # 5. Flat BI export
    flat_export = create_flat_bi_export(data_with_flags)

    # 6. Валідація
    all_passed, messages = validate_results(
        raw_data, drug_stats, drug_dist, flat_export,
        researched_drugs, total_files
    )

    # 7. Експорт CSV
    export_to_csv(drug_stats, drug_dist, flat_export, OUTPUT_BASE_PATH)

    # 8. Експорт XLSX
    export_to_xlsx(drug_stats, drug_dist, OUTPUT_BUSINESS_PATH)

    # 9. Validation report
    create_validation_report(all_passed, messages, OUTPUT_BASE_PATH)

    # Summary
    print("\n" + "=" * 70)
    print("SUMMARY")
    print("=" * 70)
    print(f"  Input: {total_files} sub_coef files, {len(raw_data)} records")
    print(f"  Drug statistics: {len(drug_stats)} drugs")
    print(f"  Drug distribution: {len(drug_dist)} drugs x {len(BIN_LABELS)} bins")
    print(f"  Flat BI export: {len(flat_export)} records")
    print(f"  Outliers removed: {data_with_flags['IS_OUTLIER'].sum()}")
    print(f"  Validation: {'PASSED' if all_passed else 'FAILED'}")
    print(f"\nOutput folder: {OUTPUT_BASE_PATH}")
    print(f"Business reports: {OUTPUT_BUSINESS_PATH}")
    print(f"\nFinished: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}")
    print("=" * 70)


if __name__ == "__main__":
    main()
