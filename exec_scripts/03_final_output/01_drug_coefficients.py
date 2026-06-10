# =============================================================================
# 01_DRUG_COEFFICIENTS - cross_pharm_market_analysis / Phase 3
# =============================================================================
# Файл: exec_scripts/03_final_output/01_drug_coefficients.py
# Дата: 2026-04-27
# Опис: Phase 3, Step 1 — Фінальні коефіцієнти субституції препаратів.
#        Фільтрація, бізнес-звіти.
# =============================================================================
"""
Формує фінальні коефіцієнти субституції на основі результатів Phase 2.

Логіка:
    1. Завантажити drug_statistics.csv (Phase 2 Step 2.1) — медіана вже розрахована
       після IQR-фільтрації outliers (поле MEDIAN_SHARE_INTERNAL)
    2. Фільтр прийнятих: MARKET_COUNT_TOTAL >= MIN_MARKET_COUNT
    3. Зберегти CSV + XLSX бізнес-звіти

Методологічна примітка:
    MEDIAN_SUBSTITUTION_COEF = MEDIAN_SHARE_INTERNAL з drug_statistics.csv.
    Цей показник розраховано ПІСЛЯ IQR-видалення outliers (Phase 2 Step 2.1),
    що забезпечує методологічну узгодженість між Phase 2 та Phase 3.
    Використання сирої матриці researched_drugs_coefficients.csv давало б
    систематичне зміщення для препаратів з outliers.

Вхід:
    results/substitution_research/02_statistics_and_filter/
        02_01_statistical_analysis/drug_statistics.csv

Вихід:
    results/substitution_research/03_final_output/03_01_drug_coefficients/
        drug_coefficients.csv
        rejected_drugs.csv
        filter_summary.csv
        validation_report.txt
        coef_business_reports/
            drug_coefficients.xlsx
            rejected_drugs.xlsx

Використання:
    python exec_scripts/03_final_output/01_drug_coefficients.py
"""

import sys
from pathlib import Path
from datetime import datetime
from typing import List, Tuple

import pandas as pd
import numpy as np

SCRIPT_PATH = Path(__file__).resolve()
PROJECT_ROOT = SCRIPT_PATH.parent.parent.parent
sys.path.insert(0, str(PROJECT_ROOT))

from project_core.data_config.paths_config import RESULTS_PATH


# =============================================================================
# КОНСТАНТИ
# =============================================================================

# Мінімальна кількість ринків (критерій Sequential Analyzer, Study 02)
MIN_MARKET_COUNT = 20

INPUT_DRUG_STATISTICS = (
    RESULTS_PATH / "substitution_research" / "02_statistics_and_filter"
    / "02_01_statistical_analysis" / "drug_statistics.csv"
)

OUTPUT_DIR = RESULTS_PATH / "substitution_research" / "03_final_output" / "03_01_drug_coefficients"
OUTPUT_BUSINESS_DIR = OUTPUT_DIR / "coef_business_reports"

OUTPUT_DRUG_COEF   = OUTPUT_DIR / "drug_coefficients.csv"
OUTPUT_REJECTED    = OUTPUT_DIR / "rejected_drugs.csv"
OUTPUT_FILTER_SUM  = OUTPUT_DIR / "filter_summary.csv"
OUTPUT_VALIDATION  = OUTPUT_DIR / "validation_report.txt"

DRUG_COEF_COLUMNS = [
    "DRUGS_ID", "DRUGS_NAME", "INN_ID", "INN_NAME",
    "NFC1_ID", "MARKET_COUNT", "MEDIAN_SUBSTITUTION_COEF",
]

REJECTED_COLUMNS = DRUG_COEF_COLUMNS + ["REJECT_REASON"]

# Порогові значення кольорового маркування коефіцієнту
COEF_HIGH_THRESHOLD   = 0.70  # >= 0.70 → green
COEF_MEDIUM_THRESHOLD = 0.40  # 0.40-0.70 → yellow, < 0.40 → red


# =============================================================================
# ЗАВАНТАЖЕННЯ
# =============================================================================

def load_drug_statistics() -> pd.DataFrame:
    """
    Завантажити drug_statistics.csv (Phase 2 Step 2.1).

    Використовує MEDIAN_SHARE_INTERNAL — медіану після IQR-фільтрації outliers.
    Перейменовує поля для уніфікованого іменування Phase 3.
    """
    if not INPUT_DRUG_STATISTICS.exists():
        raise FileNotFoundError(
            f"Файл не знайдено: {INPUT_DRUG_STATISTICS}\n"
            f"Спочатку виконайте Phase 2 Step 2.1:\n"
            f"  python exec_scripts/02_substitution_coefficients/02_01_statistical_analysis.py"
        )
    print(f"  Читання: {INPUT_DRUG_STATISTICS.name}")
    df = pd.read_csv(INPUT_DRUG_STATISTICS)
    print(f"  Завантажено: {len(df)} препаратів")

    # Перейменовуємо для уніфікації
    df = df.rename(columns={
        "MARKET_COUNT_TOTAL":  "MARKET_COUNT",
        "MEDIAN_SHARE_INTERNAL": "MEDIAN_SUBSTITUTION_COEF",
    })

    missing = [c for c in ["DRUGS_ID","DRUGS_NAME","INN_ID","INN_NAME","NFC1_ID",
                            "MARKET_COUNT","MEDIAN_SUBSTITUTION_COEF"] if c not in df.columns]
    if missing:
        raise ValueError(f"Відсутні очікувані колонки: {missing}")

    print(f"  Джерело медіани: MEDIAN_SHARE_INTERNAL (після IQR-фільтрації Phase 2)")
    return df


# =============================================================================
# ФІЛЬТРАЦІЯ
# =============================================================================

def split_by_filter(df: pd.DataFrame) -> Tuple[pd.DataFrame, pd.DataFrame]:
    """
    Розбити на прийнятих (MARKET_COUNT >= MIN_MARKET_COUNT) та відхилених.

    Returns:
        (accepted_df, rejected_df)
    """
    mask_accepted = df["MARKET_COUNT"] >= MIN_MARKET_COUNT
    accepted = df[mask_accepted].copy()
    rejected = df[~mask_accepted].copy()

    rejected["REJECT_REASON"] = f"MARKET_COUNT < {MIN_MARKET_COUNT}"

    print(f"  Фільтр MARKET_COUNT >= {MIN_MARKET_COUNT}:")
    print(f"    Прийнято:    {len(accepted)} препаратів")
    print(f"    Відхилено:   {len(rejected)} препаратів")
    print(f"    Всього:      {len(df)} препаратів")

    return accepted, rejected


# =============================================================================
# FILTER SUMMARY
# =============================================================================

def create_filter_summary(
    df_all: pd.DataFrame,
    accepted: pd.DataFrame,
    rejected: pd.DataFrame,
) -> pd.DataFrame:
    """Підсумкові метрики фільтрації."""
    total = len(df_all)
    n_accepted = len(accepted)
    n_rejected = len(rejected)
    coef = accepted["MEDIAN_SUBSTITUTION_COEF"]

    rows = [
        ("TOTAL_RESEARCHED",         total,                        "Всього досліджених препаратів"),
        ("ACCEPTED",                 n_accepted,                   f"Пройшли фільтр (MARKET_COUNT >= {MIN_MARKET_COUNT})"),
        ("REJECTED",                 n_rejected,                   f"Відхилені (MARKET_COUNT < {MIN_MARKET_COUNT})"),
        ("ACCEPTANCE_RATIO",         round(n_accepted / total, 4), "Частка прийнятих"),
        ("MIN_MARKET_COUNT",         MIN_MARKET_COUNT,             "Поріг мінімальної кількості ринків"),
        ("MAX_MARKET_COUNT_ALL",     int(df_all["MARKET_COUNT"].max()), "Максимум ринків серед усіх"),
        ("MAX_MARKET_COUNT_REJECTED",int(rejected["MARKET_COUNT"].max()) if len(rejected) > 0 else 0,
                                                                   "Максимум ринків серед відхилених"),
        ("MEDIAN_COEF_ACCEPTED",     round(float(coef.median()), 4), "Медіана MEDIAN_SUBSTITUTION_COEF (прийняті)"),
        ("MEAN_COEF_ACCEPTED",       round(float(coef.mean()), 4),   "Середнє MEDIAN_SUBSTITUTION_COEF (прийняті)"),
        ("COEF_HIGH_COUNT",          int((coef >= COEF_HIGH_THRESHOLD).sum()),
                                                                   f"Коефіцієнт >= {COEF_HIGH_THRESHOLD} (HIGH)"),
        ("COEF_MEDIUM_COUNT",        int(((coef >= COEF_MEDIUM_THRESHOLD) & (coef < COEF_HIGH_THRESHOLD)).sum()),
                                                                   f"Коефіцієнт {COEF_MEDIUM_THRESHOLD}-{COEF_HIGH_THRESHOLD} (MEDIUM)"),
        ("COEF_LOW_COUNT",           int((coef < COEF_MEDIUM_THRESHOLD).sum()),
                                                                   f"Коефіцієнт < {COEF_MEDIUM_THRESHOLD} (LOW)"),
    ]

    return pd.DataFrame(rows, columns=["METRIC", "VALUE", "DESCRIPTION"])


# =============================================================================
# ВАЛІДАЦІЯ
# =============================================================================

def validate_results(
    accepted: pd.DataFrame,
    rejected: pd.DataFrame,
    df_all: pd.DataFrame,
) -> Tuple[bool, List[str]]:
    """Валідація результатів фільтрації."""
    messages = []
    all_passed = True
    check_num = 0

    def check(name: str, condition: bool, detail: str = ""):
        nonlocal all_passed, check_num
        check_num += 1
        status = "PASSED" if condition else "FAILED"
        if not condition:
            all_passed = False
        msg = f"  [{check_num:>2}] {status}: {name}"
        if detail:
            msg += f" — {detail}"
        print(msg)
        messages.append(msg)

    total = len(df_all)

    check("COMPLETENESS",
          len(accepted) + len(rejected) == total,
          f"accepted({len(accepted)}) + rejected({len(rejected)}) = {len(accepted)+len(rejected)}, expected {total}")

    out_of_range = ((accepted["MEDIAN_SUBSTITUTION_COEF"] < 0) |
                    (accepted["MEDIAN_SUBSTITUTION_COEF"] > 1)).sum()
    check("COEF_RANGE_ACCEPTED",
          out_of_range == 0,
          f"all {len(accepted)} values in [0, 1]")

    nan_count = accepted["MEDIAN_SUBSTITUTION_COEF"].isna().sum()
    check("NO_NAN_ACCEPTED",
          nan_count == 0,
          f"NaN count: {nan_count}")

    below_min = (accepted["MARKET_COUNT"] < MIN_MARKET_COUNT).sum()
    check("MIN_MARKET_COUNT_ACCEPTED",
          below_min == 0,
          f"all accepted MARKET_COUNT >= {MIN_MARKET_COUNT}")

    above_min = (rejected["MARKET_COUNT"] >= MIN_MARKET_COUNT).sum()
    check("MARKET_COUNT_REJECTED",
          above_min == 0,
          f"no rejected drug with MARKET_COUNT >= {MIN_MARKET_COUNT}")

    duplicates = accepted["DRUGS_ID"].duplicated().sum()
    check("NO_DUPLICATES_ACCEPTED",
          duplicates == 0,
          f"duplicates: {duplicates}")

    overlap = set(accepted["DRUGS_ID"]) & set(rejected["DRUGS_ID"])
    check("NO_OVERLAP",
          len(overlap) == 0,
          f"overlap: {len(overlap)}")

    print(f"\n  Загалом: {check_num} перевірок — {'ВСІ ПРОЙШЛИ' if all_passed else 'Є ПОМИЛКИ!'}")
    return all_passed, messages


# =============================================================================
# ЗБЕРЕЖЕННЯ CSV
# =============================================================================

def save_csv(accepted: pd.DataFrame, rejected: pd.DataFrame, filter_summary: pd.DataFrame) -> None:
    """Зберегти всі CSV файли."""
    OUTPUT_DIR.mkdir(parents=True, exist_ok=True)

    acc_out = (
        accepted[DRUG_COEF_COLUMNS]
        .sort_values("MEDIAN_SUBSTITUTION_COEF", ascending=False)
        .reset_index(drop=True)
    )
    acc_out.to_csv(OUTPUT_DRUG_COEF, index=False)
    print(f"  drug_coefficients.csv:  {len(acc_out)} рядків")

    rej_out = (
        rejected[REJECTED_COLUMNS]
        .sort_values("MARKET_COUNT", ascending=False)
        .reset_index(drop=True)
    )
    rej_out.to_csv(OUTPUT_REJECTED, index=False)
    print(f"  rejected_drugs.csv:     {len(rej_out)} рядків")

    filter_summary.to_csv(OUTPUT_FILTER_SUM, index=False)
    print(f"  filter_summary.csv:     {len(filter_summary)} рядків")


# =============================================================================
# XLSX БІЗНЕС-ЗВІТИ
# =============================================================================

def export_xlsx(accepted: pd.DataFrame, rejected: pd.DataFrame) -> None:
    """Зберегти XLSX звіти з форматуванням."""
    OUTPUT_BUSINESS_DIR.mkdir(parents=True, exist_ok=True)

    try:
        from openpyxl import load_workbook
        from openpyxl.styles import PatternFill, Font, Alignment, Border, Side

        fills = {
            "HIGH":   PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid"),
            "MEDIUM": PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid"),
            "LOW":    PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid"),
        }
        bold_font  = Font(bold=True)
        header_fill = PatternFill(start_color="BDD7EE", end_color="BDD7EE", fill_type="solid")
        thin_border = Border(
            left=Side(style="thin"), right=Side(style="thin"),
            top=Side(style="thin"), bottom=Side(style="thin"),
        )

        def _apply_header(ws, n_cols: int):
            for col in range(1, n_cols + 1):
                cell = ws.cell(row=1, column=col)
                cell.font = bold_font
                cell.fill = header_fill
                cell.alignment = Alignment(horizontal="center")
                cell.border = thin_border

        def _add_borders(ws):
            for row in ws.iter_rows():
                for cell in row:
                    cell.border = thin_border

        # --- drug_coefficients.xlsx ---
        acc_out = (
            accepted[DRUG_COEF_COLUMNS]
            .sort_values("MEDIAN_SUBSTITUTION_COEF", ascending=False)
            .reset_index(drop=True)
        )
        xlsx_path = OUTPUT_BUSINESS_DIR / "drug_coefficients.xlsx"
        acc_out.to_excel(xlsx_path, index=False, sheet_name="Drug Coefficients")

        wb = load_workbook(xlsx_path)
        ws = wb.active
        cols = list(acc_out.columns)

        _apply_header(ws, len(cols))

        coef_idx = cols.index("MEDIAN_SUBSTITUTION_COEF") + 1
        for row in range(2, ws.max_row + 1):
            cell = ws.cell(row=row, column=coef_idx)
            if cell.value is not None:
                cell.number_format = "0.00%"
                v = float(cell.value)
                if v >= COEF_HIGH_THRESHOLD:
                    cell.fill = fills["HIGH"]
                elif v >= COEF_MEDIUM_THRESHOLD:
                    cell.fill = fills["MEDIUM"]
                else:
                    cell.fill = fills["LOW"]

        _add_borders(ws)
        wb.save(xlsx_path)
        print(f"  drug_coefficients.xlsx: {len(acc_out)} рядків (% формат, кольорове кодування)")

        # --- rejected_drugs.xlsx ---
        rej_out = (
            rejected[REJECTED_COLUMNS]
            .sort_values("MARKET_COUNT", ascending=False)
            .reset_index(drop=True)
        )
        xlsx_rej = OUTPUT_BUSINESS_DIR / "rejected_drugs.xlsx"
        rej_out.to_excel(xlsx_rej, index=False, sheet_name="Rejected Drugs")

        wb = load_workbook(xlsx_rej)
        ws = wb.active
        _apply_header(ws, len(rej_out.columns))

        coef_idx = list(rej_out.columns).index("MEDIAN_SUBSTITUTION_COEF") + 1
        for row in range(2, ws.max_row + 1):
            cell = ws.cell(row=row, column=coef_idx)
            if cell.value is not None:
                cell.number_format = "0.00%"

        _add_borders(ws)
        wb.save(xlsx_rej)
        print(f"  rejected_drugs.xlsx:    {len(rej_out)} рядків")

    except ImportError:
        print("  ПОМИЛКА: openpyxl не встановлено (pip install openpyxl)")
    except Exception as e:
        print(f"  ПОМИЛКА при експорті XLSX: {e}")
        import traceback
        traceback.print_exc()


# =============================================================================
# VALIDATION REPORT
# =============================================================================

def save_validation_report(all_passed: bool, messages: List[str]) -> None:
    """Зберегти текстовий звіт валідації."""
    with open(OUTPUT_VALIDATION, "w", encoding="utf-8") as f:
        f.write("=" * 70 + "\n")
        f.write("VALIDATION REPORT — Phase 3 Step 1: Drug Coefficients\n")
        f.write(f"Generated: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}\n")
        f.write("=" * 70 + "\n\n")
        f.write(f"Filter: MARKET_COUNT >= {MIN_MARKET_COUNT}\n\n")
        f.write("Validation checks:\n")
        for msg in messages:
            f.write(msg + "\n")
        f.write(f"\nOverall: {'ALL PASSED' if all_passed else 'SOME FAILED'}\n")
    print(f"  validation_report.txt:  saved")


# =============================================================================
# SUMMARY
# =============================================================================

def print_summary(accepted: pd.DataFrame) -> None:
    """Вивести підсумкову статистику."""
    coef = accepted["MEDIAN_SUBSTITUTION_COEF"]
    print(f"\n  Статистика MEDIAN_SUBSTITUTION_COEF (прийняті):")
    print(f"    Медіана:  {coef.median():.4f}")
    print(f"    Середнє: {coef.mean():.4f}")
    print(f"    Мін:     {coef.min():.4f}")
    print(f"    Макс:    {coef.max():.4f}")

    n = len(accepted)
    high   = (coef >= COEF_HIGH_THRESHOLD).sum()
    medium = ((coef >= COEF_MEDIUM_THRESHOLD) & (coef < COEF_HIGH_THRESHOLD)).sum()
    low    = (coef < COEF_MEDIUM_THRESHOLD).sum()
    print(f"\n  Розподіл по рівнях (% від прийнятих):")
    print(f"    HIGH   (>= {COEF_HIGH_THRESHOLD:.0%}): {high:>4} ({high/n:.1%})")
    print(f"    MEDIUM ({COEF_MEDIUM_THRESHOLD:.0%}-{COEF_HIGH_THRESHOLD:.0%}):   {medium:>4} ({medium/n:.1%})")
    print(f"    LOW    (< {COEF_MEDIUM_THRESHOLD:.0%}):  {low:>4} ({low/n:.1%})")


# =============================================================================
# MAIN
# =============================================================================

def main():
    print("=" * 70)
    print("PHASE 3, STEP 1: DRUG COEFFICIENTS")
    print(f"Started: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}")
    print("=" * 70)

    print("\n[1/4] Завантаження даних (Phase 2 drug_statistics)...")
    df = load_drug_statistics()

    print("\n[2/4] Фільтрація...")
    accepted, rejected = split_by_filter(df)

    print("\n[3/4] Валідація...")
    all_passed, messages = validate_results(accepted, rejected, df)

    print("\n[4/4] Збереження результатів...")
    filter_summary = create_filter_summary(df, accepted, rejected)
    save_csv(accepted, rejected, filter_summary)
    export_xlsx(accepted, rejected)
    save_validation_report(all_passed, messages)
    print_summary(accepted)

    print("\n" + "=" * 70)
    status = "PASSED" if all_passed else "FAILED (перевірте validation_report.txt)"
    print(f"РЕЗУЛЬТАТ: {status}")
    print(f"Вихідна папка: {OUTPUT_DIR}")
    print(f"Бізнес-звіти:  {OUTPUT_BUSINESS_DIR}")
    print(f"Finished: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}")
    print("=" * 70)


if __name__ == "__main__":
    main()
