# =============================================================================
# 02_SUBSTITUTE_SHARES - cross_pharm_market_analysis / Phase 3
# =============================================================================
# Файл: exec_scripts/03_final_output/02_substitute_shares.py
# Дата: 2026-04-27
# Опис: Phase 3, Step 2 — Крос-ринкова агрегація субститутів (LIFT-зважена).
# =============================================================================
"""
Агрегує дані субститутів з усіх ринків у єдиний фінальний файл.

Метод: LIFT-зважена агрегація (аналогічно WEIGHTED_MEAN_SHARE Phase 2).
    Для кожного ринку:
        TOTAL_LIFT = SUBSTITUTE_SHARE (decimal) × INTERNAL_LIFT
    Cross-market:
        AGG_SHARE = SUM(TOTAL_LIFT) / SUM(INTERNAL_LIFT per market per drug)
    Інваріант: SUM(AGG_SHARE per stockout drug) = 1.0

Вхід:
    results/cross_market_data/market_substitution_{ID}/sub_drugs_{ID}.csv
    results/cross_market_data/market_substitution_{ID}/sub_coef_{ID}.csv
    results/substitution_research/03_final_output/03_01_drug_coefficients/drug_coefficients.csv

Вихід:
    results/substitution_research/03_final_output/03_02_substitute_shares/
        substitute_shares.csv
        substitute_summary.csv
        validation_report.txt
        subst_business_reports/
            substitute_shares.xlsx
            substitute_summary.xlsx

Використання:
    python exec_scripts/03_final_output/02_substitute_shares.py
"""

import sys
from pathlib import Path
from datetime import datetime
from typing import List, Tuple

import pandas as pd
import numpy as np

SCRIPT_PATH = Path(__file__).resolve()
PROJECT_ROOT = SCRIPT_PATH.parent.parent.parent
sys.path.insert(0, str(PROJECT_ROOT))

from project_core.data_config.paths_config import RESULTS_PATH


# =============================================================================
# КОНСТАНТИ
# =============================================================================

CROSS_MARKET_PATH = RESULTS_PATH / "cross_market_data"

DRUG_COEF_PATH = (
    RESULTS_PATH / "substitution_research" / "03_final_output"
    / "03_01_drug_coefficients" / "drug_coefficients.csv"
)

OUTPUT_DIR = (
    RESULTS_PATH / "substitution_research" / "03_final_output"
    / "03_02_substitute_shares"
)
OUTPUT_BUSINESS_DIR = OUTPUT_DIR / "subst_business_reports"

OUTPUT_SHARES     = OUTPUT_DIR / "substitute_shares.csv"
OUTPUT_SUMMARY    = OUTPUT_DIR / "substitute_summary.csv"
OUTPUT_VALIDATION = OUTPUT_DIR / "validation_report.txt"

# Допустима похибка float для перевірки суми часток
SHARE_SUM_EPSILON = 0.01

SUB_COEF_COLS = ["CLIENT_ID", "DRUGS_ID", "INTERNAL_LIFT"]
SUB_DRUGS_COLS = [
    "CLIENT_ID", "STOCKOUT_DRUG_ID", "STOCKOUT_DRUG_NAME",
    "INN_ID", "INN_NAME", "NFC1_ID",
    "SUBSTITUTE_DRUG_ID", "SUBSTITUTE_DRUG_NAME", "SUBSTITUTE_NFC1_ID",
    "SAME_NFC1", "SUBSTITUTE_SHARE",
]

SHARES_COLUMNS = [
    "STOCKOUT_DRUG_ID", "STOCKOUT_DRUG_NAME",
    "INN_ID", "INN_NAME", "NFC1_ID",
    "SUBSTITUTE_DRUG_ID", "SUBSTITUTE_DRUG_NAME", "SUBSTITUTE_NFC1_ID",
    "SAME_NFC1", "AGG_SUBSTITUTE_SHARE", "MARKETS_COUNT", "SUBSTITUTE_RANK",
]

SUMMARY_COLUMNS = [
    "STOCKOUT_DRUG_ID", "STOCKOUT_DRUG_NAME",
    "INN_ID", "INN_NAME", "NFC1_ID",
    "N_SUBSTITUTES",
    "TOP_SUBSTITUTE_ID", "TOP_SUBSTITUTE_NAME", "TOP_SUBSTITUTE_SHARE",
    "TOTAL_MARKETS_COVERED", "AVG_MARKETS_PER_SUBSTITUTE",
]


# =============================================================================
# ЗАВАНТАЖЕННЯ
# =============================================================================

def load_accepted_drug_ids() -> set:
    """Завантажити список прийнятих препаратів з Step 1."""
    if not DRUG_COEF_PATH.exists():
        raise FileNotFoundError(
            f"drug_coefficients.csv не знайдено: {DRUG_COEF_PATH}\n"
            f"Спочатку виконайте Step 1:\n"
            f"  python exec_scripts/03_final_output/01_drug_coefficients.py"
        )
    df = pd.read_csv(DRUG_COEF_PATH, usecols=["DRUGS_ID"])
    ids = set(df["DRUGS_ID"].tolist())
    print(f"  Завантажено препаратів з Step 1: {len(ids)}")
    return ids


def find_market_folders() -> List[Path]:
    """Знайти всі папки market_substitution_*."""
    folders = sorted(CROSS_MARKET_PATH.glob("market_substitution_*/"))
    if not folders:
        raise FileNotFoundError(
            f"Не знайдено папок market_substitution_* у {CROSS_MARKET_PATH}"
        )
    return folders


def load_market_pair(folder: Path, market_id: int) -> Tuple[pd.DataFrame, pd.DataFrame]:
    """Завантажити sub_drugs та sub_coef для одного ринку."""
    sub_drugs_file = folder / f"sub_drugs_{market_id}.csv"
    sub_coef_file  = folder / f"sub_coef_{market_id}.csv"
    empty = pd.DataFrame()

    if not sub_drugs_file.exists() or not sub_coef_file.exists():
        return empty, empty

    sub_drugs = pd.read_csv(sub_drugs_file, usecols=SUB_DRUGS_COLS, dtype={"SAME_NFC1": str})
    sub_coef  = pd.read_csv(sub_coef_file,  usecols=SUB_COEF_COLS)
    return sub_drugs, sub_coef


# =============================================================================
# ОБРОБКА ОДНОГО РИНКУ
# =============================================================================

def process_market(
    sub_drugs: pd.DataFrame,
    sub_coef: pd.DataFrame,
    valid_drug_ids: set,
    market_id: int,
) -> pd.DataFrame:
    """
    Розрахувати TOTAL_LIFT для кожної пари (stockout, substitute) на одному ринку.
    TOTAL_LIFT = SUBSTITUTE_SHARE (decimal 0-1) × INTERNAL_LIFT
    """
    if len(sub_drugs) == 0 or len(sub_coef) == 0:
        return pd.DataFrame()

    valid_subs = sub_drugs[
        sub_drugs["SUBSTITUTE_DRUG_ID"].notna() &
        (sub_drugs["SUBSTITUTE_DRUG_ID"] != "") &
        sub_drugs["SUBSTITUTE_SHARE"].notna() &
        (sub_drugs["SUBSTITUTE_SHARE"] != "")
    ].copy()

    if len(valid_subs) == 0:
        return pd.DataFrame()

    valid_subs["SUBSTITUTE_SHARE"]  = pd.to_numeric(valid_subs["SUBSTITUTE_SHARE"],  errors="coerce")
    valid_subs["STOCKOUT_DRUG_ID"]  = pd.to_numeric(valid_subs["STOCKOUT_DRUG_ID"],  errors="coerce")
    valid_subs["SUBSTITUTE_DRUG_ID"]= pd.to_numeric(valid_subs["SUBSTITUTE_DRUG_ID"],errors="coerce")

    valid_subs = valid_subs[valid_subs["STOCKOUT_DRUG_ID"].isin(valid_drug_ids)]
    if len(valid_subs) == 0:
        return pd.DataFrame()

    coef_filtered = sub_coef[
        sub_coef["DRUGS_ID"].isin(valid_drug_ids) &
        (sub_coef["INTERNAL_LIFT"] > 0)
    ][["DRUGS_ID", "INTERNAL_LIFT"]].copy()
    coef_filtered.columns = ["STOCKOUT_DRUG_ID", "INTERNAL_LIFT"]

    result = valid_subs.merge(coef_filtered, on="STOCKOUT_DRUG_ID", how="inner")
    if len(result) == 0:
        return pd.DataFrame()

    result["TOTAL_LIFT"] = result["SUBSTITUTE_SHARE"] * result["INTERNAL_LIFT"]
    result = result[result["TOTAL_LIFT"] > 0].copy()
    result["CLIENT_ID"] = market_id

    return result


# =============================================================================
# КРОС-РИНКОВА АГРЕГАЦІЯ
# =============================================================================

def aggregate_cross_market(all_market_data: List[pd.DataFrame]) -> pd.DataFrame:
    """
    AGG_SHARE = SUM(TOTAL_LIFT per pair) / SUM(INTERNAL_LIFT per market per drug).
    INTERNAL_LIFT дедублікується по (CLIENT_ID, STOCKOUT_DRUG_ID).
    """
    print(f"\n  Об'єднання даних з {len(all_market_data)} ринків...")
    combined = pd.concat(all_market_data, ignore_index=True)
    print(f"  Всього записів: {len(combined):,}")
    print(f"  Унікальних STOCKOUT_DRUG_ID:   {combined['STOCKOUT_DRUG_ID'].nunique()}")
    print(f"  Унікальних SUBSTITUTE_DRUG_ID: {combined['SUBSTITUTE_DRUG_ID'].nunique()}")

    pair_meta_cols = [
        "STOCKOUT_DRUG_ID", "STOCKOUT_DRUG_NAME",
        "INN_ID", "INN_NAME", "NFC1_ID",
        "SUBSTITUTE_DRUG_ID", "SUBSTITUTE_DRUG_NAME", "SUBSTITUTE_NFC1_ID", "SAME_NFC1",
    ]

    agg_total_lift = combined.groupby(pair_meta_cols, as_index=False).agg(
        TOTAL_LIFT_SUM=("TOTAL_LIFT",  "sum"),
        MARKETS_COUNT=("TOTAL_LIFT", "count"),
    )

    # INTERNAL_LIFT дедублікується — одне значення на (ринок, stockout_drug)
    drug_lift = (
        combined
        .drop_duplicates(subset=["CLIENT_ID", "STOCKOUT_DRUG_ID"])
        .groupby("STOCKOUT_DRUG_ID", as_index=False)["INTERNAL_LIFT"]
        .sum()
        .rename(columns={"INTERNAL_LIFT": "INTERNAL_LIFT_TOTAL"})
    )

    result = agg_total_lift.merge(drug_lift, on="STOCKOUT_DRUG_ID", how="left")
    result["AGG_SUBSTITUTE_SHARE"] = np.where(
        result["INTERNAL_LIFT_TOTAL"] > 0,
        result["TOTAL_LIFT_SUM"] / result["INTERNAL_LIFT_TOTAL"],
        0.0,
    )

    result = result.sort_values(
        ["STOCKOUT_DRUG_ID", "AGG_SUBSTITUTE_SHARE"],
        ascending=[True, False],
    ).reset_index(drop=True)

    result["SUBSTITUTE_RANK"] = result.groupby("STOCKOUT_DRUG_ID").cumcount() + 1

    return result


# =============================================================================
# SUBSTITUTE SUMMARY
# =============================================================================

def create_substitute_summary(df: pd.DataFrame) -> pd.DataFrame:
    """
    Per-drug підсумок: кількість субститутів, топ-субститут, покриття ринків.
    """
    rows = []
    for drug_id, group in df.groupby("STOCKOUT_DRUG_ID"):
        group_sorted = group.sort_values("SUBSTITUTE_RANK")
        top = group_sorted.iloc[0]
        rows.append({
            "STOCKOUT_DRUG_ID":          drug_id,
            "STOCKOUT_DRUG_NAME":        top["STOCKOUT_DRUG_NAME"],
            "INN_ID":                    top["INN_ID"],
            "INN_NAME":                  top["INN_NAME"],
            "NFC1_ID":                   top["NFC1_ID"],
            "N_SUBSTITUTES":             len(group),
            "TOP_SUBSTITUTE_ID":         top["SUBSTITUTE_DRUG_ID"],
            "TOP_SUBSTITUTE_NAME":       top["SUBSTITUTE_DRUG_NAME"],
            "TOP_SUBSTITUTE_SHARE":      round(float(top["AGG_SUBSTITUTE_SHARE"]), 6),
            "TOTAL_MARKETS_COVERED":     int(group["MARKETS_COUNT"].max()),
            "AVG_MARKETS_PER_SUBSTITUTE":round(float(group["MARKETS_COUNT"].mean()), 1),
        })
    summary = pd.DataFrame(rows, columns=SUMMARY_COLUMNS)
    return summary.sort_values("N_SUBSTITUTES", ascending=False).reset_index(drop=True)


# =============================================================================
# ВАЛІДАЦІЯ
# =============================================================================

def validate_results(df: pd.DataFrame) -> Tuple[bool, List[str]]:
    """Валідація фінального результату."""
    messages = []
    all_passed = True
    check_num = 0

    def check(name: str, condition: bool, detail: str = ""):
        nonlocal all_passed, check_num
        check_num += 1
        status = "PASSED" if condition else "FAILED"
        if not condition:
            all_passed = False
        msg = f"  [{check_num:>2}] {status}: {name}"
        if detail:
            msg += f" — {detail}"
        print(msg)
        messages.append(msg)

    out_of_range = ((df["AGG_SUBSTITUTE_SHARE"] < 0) |
                    (df["AGG_SUBSTITUTE_SHARE"] > 1)).sum()
    check("SHARE_RANGE",
          out_of_range == 0,
          f"all {len(df)} values in [0, 1]")

    share_sums = df.groupby("STOCKOUT_DRUG_ID")["AGG_SUBSTITUTE_SHARE"].sum()
    deviations = share_sums[abs(share_sums - 1.0) > SHARE_SUM_EPSILON]
    check("SHARE_SUM_INVARIANT",
          len(deviations) == 0,
          f"{len(deviations)} drugs with sum != 1.0 (±{SHARE_SUM_EPSILON})")
    if len(deviations) > 0:
        for did, total in deviations.head(5).items():
            print(f"      DRUGS_ID={did}: sum={total:.6f}")

    for col in ["STOCKOUT_DRUG_ID", "SUBSTITUTE_DRUG_ID", "AGG_SUBSTITUTE_SHARE"]:
        nan_count = df[col].isna().sum()
        check(f"NO_NAN_{col}",
              nan_count == 0,
              f"NaN: {nan_count}")

    duplicates = df.duplicated(subset=["STOCKOUT_DRUG_ID", "SUBSTITUTE_DRUG_ID"]).sum()
    check("NO_DUPLICATE_PAIRS",
          duplicates == 0,
          f"duplicate pairs: {duplicates}")

    rank1_count = (df["SUBSTITUTE_RANK"] == 1).sum()
    n_drugs = df["STOCKOUT_DRUG_ID"].nunique()
    check("RANK1_COUNT",
          rank1_count == n_drugs,
          f"rank-1 count={rank1_count}, unique drugs={n_drugs}")

    print(f"\n  Загалом: {check_num} перевірок — {'ВСІ ПРОЙШЛИ' if all_passed else 'Є ПОМИЛКИ!'}")
    return all_passed, messages


# =============================================================================
# ЗБЕРЕЖЕННЯ CSV
# =============================================================================

def save_csv(shares: pd.DataFrame, summary: pd.DataFrame) -> None:
    """Зберегти CSV файли."""
    OUTPUT_DIR.mkdir(parents=True, exist_ok=True)

    shares[SHARES_COLUMNS].to_csv(OUTPUT_SHARES, index=False)
    print(f"  substitute_shares.csv:  {len(shares)} рядків")

    summary.to_csv(OUTPUT_SUMMARY, index=False)
    print(f"  substitute_summary.csv: {len(summary)} рядків")


# =============================================================================
# XLSX БІЗНЕС-ЗВІТИ
# =============================================================================

def export_xlsx(shares: pd.DataFrame, summary: pd.DataFrame) -> None:
    """Зберегти XLSX бізнес-звіти з форматуванням."""
    OUTPUT_BUSINESS_DIR.mkdir(parents=True, exist_ok=True)

    try:
        from openpyxl import load_workbook
        from openpyxl.styles import PatternFill, Font, Alignment, Border, Side

        bold_font   = Font(bold=True)
        header_fill = PatternFill(start_color="BDD7EE", end_color="BDD7EE", fill_type="solid")
        rank1_fill  = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
        thin_border = Border(
            left=Side(style="thin"), right=Side(style="thin"),
            top=Side(style="thin"), bottom=Side(style="thin"),
        )

        def _apply_header(ws, n_cols: int):
            for col in range(1, n_cols + 1):
                cell = ws.cell(row=1, column=col)
                cell.font = bold_font
                cell.fill = header_fill
                cell.alignment = Alignment(horizontal="center")
                cell.border = thin_border

        def _add_borders(ws):
            for row in ws.iter_rows():
                for cell in row:
                    cell.border = thin_border

        # --- substitute_shares.xlsx ---
        shares_out = shares[SHARES_COLUMNS].copy()
        xlsx_shares = OUTPUT_BUSINESS_DIR / "substitute_shares.xlsx"
        shares_out.to_excel(xlsx_shares, index=False, sheet_name="Substitute Shares")

        wb = load_workbook(xlsx_shares)
        ws = wb.active
        cols = list(shares_out.columns)
        _apply_header(ws, len(cols))

        share_idx = cols.index("AGG_SUBSTITUTE_SHARE") + 1
        rank_idx  = cols.index("SUBSTITUTE_RANK") + 1

        for row in range(2, ws.max_row + 1):
            share_cell = ws.cell(row=row, column=share_idx)
            rank_cell  = ws.cell(row=row, column=rank_idx)
            if share_cell.value is not None:
                share_cell.number_format = "0.00%"
            if rank_cell.value == 1:
                for col in range(1, len(cols) + 1):
                    ws.cell(row=row, column=col).fill = rank1_fill

        _add_borders(ws)
        wb.save(xlsx_shares)
        print(f"  substitute_shares.xlsx: {len(shares_out)} рядків (% формат, rank-1 зелений)")

        # --- substitute_summary.xlsx ---
        xlsx_summary = OUTPUT_BUSINESS_DIR / "substitute_summary.xlsx"
        summary.to_excel(xlsx_summary, index=False, sheet_name="Summary")

        wb = load_workbook(xlsx_summary)
        ws = wb.active
        cols_s = list(summary.columns)
        _apply_header(ws, len(cols_s))

        top_share_idx = cols_s.index("TOP_SUBSTITUTE_SHARE") + 1
        for row in range(2, ws.max_row + 1):
            cell = ws.cell(row=row, column=top_share_idx)
            if cell.value is not None:
                cell.number_format = "0.00%"

        _add_borders(ws)
        wb.save(xlsx_summary)
        print(f"  substitute_summary.xlsx:{len(summary)} рядків")

    except ImportError:
        print("  ПОМИЛКА: openpyxl не встановлено (pip install openpyxl)")
    except Exception as e:
        print(f"  ПОМИЛКА при експорті XLSX: {e}")
        import traceback
        traceback.print_exc()


# =============================================================================
# VALIDATION REPORT
# =============================================================================

def save_validation_report(all_passed: bool, messages: List[str]) -> None:
    """Зберегти текстовий звіт валідації."""
    with open(OUTPUT_VALIDATION, "w", encoding="utf-8") as f:
        f.write("=" * 70 + "\n")
        f.write("VALIDATION REPORT — Phase 3 Step 2: Substitute Shares\n")
        f.write(f"Generated: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}\n")
        f.write("=" * 70 + "\n\n")
        f.write("Method: LIFT-weighted aggregation\n")
        f.write("  AGG_SHARE = SUM(TOTAL_LIFT) / SUM(INTERNAL_LIFT per market per drug)\n\n")
        f.write("Validation checks:\n")
        for msg in messages:
            f.write(msg + "\n")
        f.write(f"\nOverall: {'ALL PASSED' if all_passed else 'SOME FAILED'}\n")
    print(f"  validation_report.txt:  saved")


# =============================================================================
# SUMMARY
# =============================================================================

def print_summary(shares: pd.DataFrame) -> None:
    """Вивести підсумкову статистику."""
    n_drugs = shares["STOCKOUT_DRUG_ID"].nunique()
    n_subs  = shares["SUBSTITUTE_DRUG_ID"].nunique()
    n_pairs = len(shares)

    print(f"\n  Підсумок:")
    print(f"    Досліджуваних препаратів:  {n_drugs}")
    print(f"    Унікальних субститутів:    {n_subs}")
    print(f"    Пар (stockout→substitute): {n_pairs}")

    subs_per_drug = shares.groupby("STOCKOUT_DRUG_ID")["SUBSTITUTE_DRUG_ID"].count()
    print(f"\n  Субститутів per препарат:")
    print(f"    Медіана: {subs_per_drug.median():.0f}")
    print(f"    Макс:    {subs_per_drug.max()}")
    print(f"    1:       {(subs_per_drug == 1).sum()} препаратів")
    print(f"    2-5:     {((subs_per_drug >= 2) & (subs_per_drug <= 5)).sum()} препаратів")
    print(f"    6+:      {(subs_per_drug >= 6).sum()} препаратів")

    if "SAME_NFC1" in shares.columns:
        same_count = (shares["SAME_NFC1"].astype(str).str.upper() == "TRUE").sum()
        print(f"\n  Тип субституту:")
        print(f"    SAME_NFC1:  {same_count} пар ({same_count/n_pairs*100:.1f}%)")
        print(f"    DIFF_NFC1:  {n_pairs - same_count} пар ({(n_pairs - same_count)/n_pairs*100:.1f}%)")


# =============================================================================
# MAIN
# =============================================================================

def main():
    print("=" * 70)
    print("PHASE 3, STEP 2: SUBSTITUTE SHARES (LIFT-WEIGHTED)")
    print(f"Started: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}")
    print("=" * 70)

    print("\n[1/5] Завантаження списку прийнятих препаратів (Step 1)...")
    valid_drug_ids = load_accepted_drug_ids()

    print("\n[2/5] Читання даних по ринках...")
    market_folders = find_market_folders()
    print(f"  Знайдено папок: {len(market_folders)}")

    all_market_data = []
    skipped = 0
    processed = 0

    for folder in market_folders:
        try:
            market_id = int(folder.name.replace("market_substitution_", ""))
        except ValueError:
            continue

        sub_drugs, sub_coef = load_market_pair(folder, market_id)
        if len(sub_drugs) == 0:
            skipped += 1
            continue

        market_result = process_market(sub_drugs, sub_coef, valid_drug_ids, market_id)

        if len(market_result) > 0:
            all_market_data.append(market_result)
            processed += 1
        else:
            skipped += 1

        if processed % 20 == 0 or processed == 1:
            print(f"  [{processed}/{len(market_folders)}] market {market_id}: "
                  f"{len(market_result)} пар")

    print(f"\n  Успішно оброблено: {processed}")
    print(f"  Пропущено:         {skipped}")

    if not all_market_data:
        print("\nПОМИЛКА: Немає даних для агрегації")
        return

    print("\n[3/5] Крос-ринкова LIFT-зважена агрегація...")
    shares = aggregate_cross_market(all_market_data)
    summary = create_substitute_summary(shares)

    print("\n[4/5] Валідація...")
    all_passed, messages = validate_results(shares)

    print("\n[5/5] Збереження результатів...")
    save_csv(shares, summary)
    export_xlsx(shares, summary)
    save_validation_report(all_passed, messages)
    print_summary(shares)

    print("\n" + "=" * 70)
    status = "PASSED" if all_passed else "FAILED (перевірте validation_report.txt)"
    print(f"РЕЗУЛЬТАТ: {status}")
    print(f"Вихідна папка: {OUTPUT_DIR}")
    print(f"Бізнес-звіти:  {OUTPUT_BUSINESS_DIR}")
    print(f"Finished: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}")
    print("=" * 70)


if __name__ == "__main__":
    main()
